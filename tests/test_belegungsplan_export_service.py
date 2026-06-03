"""
Tests fuer BelegungsplanExportService (FA-2500).

Prueft: Excel-Struktur, Tabellenblatt-Namen, Spaltenanzahl,
Titelzeile, Datensatz-Zeilen, GA-Leer-Warnung (Rotmarkierung)
und Aktor-Tab bei leerem Aktor-Plan.
"""
import os
import sys
import pytest
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from knix_arranger.services.belegungsplan_service import (
    BelegungsplanData, SensorRow, ActorRow,
)
from knix_arranger.services.belegungsplan_export_service import BelegungsplanExportService

pytest.importorskip("openpyxl")
import openpyxl


# ── Hilfsfunktionen ────────────────────────────────────────────────────────────

def _sensor_row(**kwargs) -> SensorRow:
    defaults = dict(
        floor_name="EG", zone_name="Wohnung 1", room_number="E01",
        room_name="Wohnzimmer", sensor_type="Taster 2-fach",
        physical_address="1.1.2", taste_label="Taste 1",
        function="Licht schalten", ga_designation="L_E01_01 E/A",
        ga_address="1/0/0", dpt="DPST-1-1",
    )
    defaults.update(kwargs)
    return SensorRow(**defaults)


def _actor_row(**kwargs) -> ActorRow:
    defaults = dict(
        line_name="Linie 1", area_number=1, line_number=1,
        uv_location="UV EG", actor_type="Schaltaktor 4-fach",
        physical_address="1.1.1", channel_number="1", element_number=1,
        room_number="E01", zone_name="Wohnung 1", room_name="Wohnzimmer",
        gewerk_code="L", function_name="E/A",
        ga_designation="L_E01_01 E/A", ga_address="1/0/0", dpt="DPST-1-1",
    )
    defaults.update(kwargs)
    return ActorRow(**defaults)


def _export(data: BelegungsplanData) -> str:
    """Exportiert BelegungsplanData in eine temporaere Datei und gibt den Pfad zurueck."""
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    BelegungsplanExportService().export_xlsx(data, path)
    return path


# ── Tests: ZIP / Workbook-Struktur ────────────────────────────────────────────

class TestWorkbookStruktur:

    def test_datei_wird_erstellt(self):
        data = BelegungsplanData(project_name="Test", sensor_rows=[], actor_rows=[])
        path = _export(data)
        assert os.path.exists(path)
        assert os.path.getsize(path) > 0
        os.unlink(path)

    def test_sensor_tab_vorhanden(self):
        data = BelegungsplanData(project_name="Test", sensor_rows=[], actor_rows=[])
        path = _export(data)
        wb = openpyxl.load_workbook(path)
        assert "Taster & Sensoren" in wb.sheetnames
        os.unlink(path)

    def test_aktor_tab_bei_leeren_aktor_rows_nicht_vorhanden(self):
        """Wenn keine Aktor-Zeilen vorhanden, wird kein Aktoren-Tab angelegt."""
        data = BelegungsplanData(project_name="Test", sensor_rows=[], actor_rows=[])
        path = _export(data)
        wb = openpyxl.load_workbook(path)
        assert "Aktoren" not in wb.sheetnames
        os.unlink(path)

    def test_aktor_tab_vorhanden_wenn_aktor_rows(self):
        data = BelegungsplanData(
            project_name="Test",
            sensor_rows=[],
            actor_rows=[_actor_row()],
        )
        path = _export(data)
        wb = openpyxl.load_workbook(path)
        assert "Aktoren" in wb.sheetnames
        os.unlink(path)


# ── Tests: Sensor-Tab Inhalt ──────────────────────────────────────────────────

class TestSensorTab:

    def test_12_header_spalten(self):
        data = BelegungsplanData(project_name="Test", sensor_rows=[], actor_rows=[])
        path = _export(data)
        ws = openpyxl.load_workbook(path)["Taster & Sensoren"]
        # Header in Zeile 3 (12 Spalten seit Einführung der "Aktion"-Spalte)
        headers = [ws.cell(row=3, column=c).value for c in range(1, 13)]
        assert len([h for h in headers if h]) == 12
        os.unlink(path)

    def test_titelzeile_enthaelt_projektname(self):
        data = BelegungsplanData(project_name="MeinProjekt", sensor_rows=[], actor_rows=[])
        path = _export(data)
        ws = openpyxl.load_workbook(path)["Taster & Sensoren"]
        assert "MeinProjekt" in str(ws["A1"].value)
        os.unlink(path)

    def test_sensor_row_wird_in_zeile_4_geschrieben(self):
        row = _sensor_row()
        data = BelegungsplanData(project_name="Test", sensor_rows=[row], actor_rows=[])
        path = _export(data)
        ws = openpyxl.load_workbook(path)["Taster & Sensoren"]
        # Zeile 4 = erste Datenzeile (1=Titel, 2=Datum, 3=Header)
        assert ws.cell(row=4, column=3).value == "E01"   # Raum-Nr. (Spalte C)
        assert ws.cell(row=4, column=11).value == "1/0/0"  # GA-Adresse (Spalte K, nach Aktion-Spalte)
        os.unlink(path)

    def test_mehrere_sensor_rows_korrekte_zeilenanzahl(self):
        rows = [_sensor_row(room_number=f"E0{i}") for i in range(5)]
        data = BelegungsplanData(project_name="Test", sensor_rows=rows, actor_rows=[])
        path = _export(data)
        ws = openpyxl.load_workbook(path)["Taster & Sensoren"]
        # Daten in Zeilen 4–8
        filled = sum(
            1 for r in range(4, 9)
            if ws.cell(row=r, column=3).value is not None
        )
        assert filled == 5
        os.unlink(path)

    def test_leere_ga_adresse_wird_rot_markiert(self):
        """Zeile mit leerem ga_address muss rote Hintergrundfarbe haben (warn_fill)."""
        row = _sensor_row(ga_address="", ga_designation="", dpt="")
        data = BelegungsplanData(project_name="Test", sensor_rows=[row], actor_rows=[])
        path = _export(data)
        ws = openpyxl.load_workbook(path, data_only=True)["Taster & Sensoren"]
        # Spalte K (11) = GA-Adresse → sollte rot hinterlegt sein
        cell_k = ws.cell(row=4, column=11)
        fill_rgb = (cell_k.fill.fgColor.rgb if cell_k.fill else None)
        assert fill_rgb is not None
        assert "FFCDD2" in str(fill_rgb).upper(), f"Erwartet FFCDD2, bekommen: {fill_rgb}"
        os.unlink(path)

    def test_gefuellte_ga_adresse_nicht_rot(self):
        row = _sensor_row(ga_address="1/0/0")
        data = BelegungsplanData(project_name="Test", sensor_rows=[row], actor_rows=[])
        path = _export(data)
        ws = openpyxl.load_workbook(path, data_only=True)["Taster & Sensoren"]
        cell_k = ws.cell(row=4, column=11)
        fill_rgb = str(cell_k.fill.fgColor.rgb if cell_k.fill else "").upper()
        assert "FFCDD2" not in fill_rgb
        os.unlink(path)


# ── Tests: Aktor-Tab Inhalt ───────────────────────────────────────────────────

class TestAktorTab:

    def test_13_header_spalten(self):
        data = BelegungsplanData(
            project_name="Test", sensor_rows=[], actor_rows=[_actor_row()]
        )
        path = _export(data)
        ws = openpyxl.load_workbook(path)["Aktoren"]
        headers = [ws.cell(row=3, column=c).value for c in range(1, 14)]
        assert len([h for h in headers if h]) == 13
        os.unlink(path)

    def test_aktor_row_wird_in_zeile_4_geschrieben(self):
        row = _actor_row()
        data = BelegungsplanData(project_name="Test", sensor_rows=[], actor_rows=[row])
        path = _export(data)
        ws = openpyxl.load_workbook(path)["Aktoren"]
        assert ws.cell(row=4, column=4).value == "1.1.1"   # Phys. Adresse (Spalte D=4)
        assert ws.cell(row=4, column=5).value == "1"       # Kanal (Spalte E=5)
        os.unlink(path)

    def test_aktor_titelzeile_enthaelt_projektname(self):
        data = BelegungsplanData(
            project_name="ProjektXY", sensor_rows=[], actor_rows=[_actor_row()]
        )
        path = _export(data)
        ws = openpyxl.load_workbook(path)["Aktoren"]
        assert "ProjektXY" in str(ws["A1"].value)
        os.unlink(path)

    def test_mehrere_aktor_rows_korrekte_zeilenanzahl(self):
        rows = [_actor_row(physical_address=f"1.1.{i}", channel_number=str(i))
                for i in range(1, 6)]
        data = BelegungsplanData(project_name="Test", sensor_rows=[], actor_rows=rows)
        path = _export(data)
        ws = openpyxl.load_workbook(path)["Aktoren"]
        filled = sum(
            1 for r in range(4, 9)
            if ws.cell(row=r, column=4).value is not None
        )
        assert filled == 5
        os.unlink(path)

    def test_linienspalte_enthaelt_bereich_und_liniennummer(self):
        """Spalte A enthält '<area>.<line> <linienname>'."""
        row = _actor_row(area_number=2, line_number=3, line_name="Steigzone")
        data = BelegungsplanData(project_name="Test", sensor_rows=[], actor_rows=[row])
        path = _export(data)
        ws = openpyxl.load_workbook(path)["Aktoren"]
        linie_val = str(ws.cell(row=4, column=1).value)
        assert "2.3" in linie_val
        assert "Steigzone" in linie_val
        os.unlink(path)


# ── Tests: Autofilter und Freeze Panes ────────────────────────────────────────

class TestExcelFeatures:

    def test_sensor_tab_hat_autofilter(self):
        data = BelegungsplanData(project_name="Test", sensor_rows=[_sensor_row()], actor_rows=[])
        path = _export(data)
        ws = openpyxl.load_workbook(path)["Taster & Sensoren"]
        assert ws.auto_filter.ref is not None
        os.unlink(path)

    def test_sensor_tab_hat_freeze_panes(self):
        data = BelegungsplanData(project_name="Test", sensor_rows=[_sensor_row()], actor_rows=[])
        path = _export(data)
        ws = openpyxl.load_workbook(path)["Taster & Sensoren"]
        assert ws.freeze_panes == "A4"
        os.unlink(path)

    def test_aktor_tab_hat_autofilter(self):
        data = BelegungsplanData(project_name="Test", sensor_rows=[], actor_rows=[_actor_row()])
        path = _export(data)
        ws = openpyxl.load_workbook(path)["Aktoren"]
        assert ws.auto_filter.ref is not None
        os.unlink(path)
