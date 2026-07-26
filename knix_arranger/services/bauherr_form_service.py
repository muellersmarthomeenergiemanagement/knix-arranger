"""
Funktionsdefinitions-Formular für Bauherren (FA-1501, FA-1502, FA-1503)
Erzeugt Excel-Formulare zur Abstimmung der Tastenbelegung mit dem Bauherren.

Grafische Taster-Darstellung (FA-1504):
Jeder Taster wird als visuelle Schaltflächen-Gruppe dargestellt.
Pro Wippe (rocker) gibt es eine obere und untere Taste, die farbig
mit ihrer Funktion beschriftet sind – so wie sie auf der Wand sitzen.

Leere Slots erhalten eine Dropdown-Auswahl mit Wunschfunktionen (Option C).
"""
from __future__ import annotations
import logging
from datetime import datetime

from ..models.project import KnxProject
from ..models.building import Room, Bedienelement, FunctionAssignment, SensorFunktion
from ..utils.excel_generator import ExcelGenerator, HAS_OPENPYXL

logger = logging.getLogger("knix_arranger.bauherr_form")

if HAS_OPENPYXL:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ── Farben für die Taster-Grafik ──────────────────────────────────────────────
_BTN_FILL_FREE     = "F5F5F5"   # Grau:       Taste ohne Funktion
_BTN_FILL_WISH     = "FFF8E1"   # Hellgelb:   Leere Taste mit Dropdown-Wunsch
_BTN_FILL_ASSIGNED = "E8F5E9"   # Hellgrün:   Taste mit eigener Funktion
_BTN_FILL_FOREIGN  = "E3F2FD"   # Hellblau:   Taste mit Fremdraum-Funktion
_BTN_FILL_GA       = "F3E5F5"   # Hellviolett: Taste mit direkter GA (Sonderwunsch)
_BTN_FILL_HEADER   = "1A5276"   # Dunkelblau: Geräte-Header
_BTN_FONT_LABEL    = "757575"   # Grau:       "oben" / "unten" Beschriftung
_BTN_FONT_FN       = "1B5E20"   # Dunkelgrün: Funktionstext
_BTN_FONT_WISH     = "E65100"   # Orange:     Wunsch-Platzhaltertext
_BTN_FONT_EMPTY    = "BDBDBD"   # Hellgrau:   Leertext

# ── Bauherren-lesbare Funktionsnamen je Gewerk + Tastenrichtung ───────────────
# (code, is_primary=True für "oben", False für "unten")
_GEWERK_BTN_LABEL: dict[tuple[str, bool], str] = {
    ("L",   True):  "Licht Ein/Aus",
    ("L",   False): "Licht Ein/Aus",
    ("LD",  True):  "Licht Ein / Heller",
    ("LD",  False): "Licht Aus / Dunkler",
    ("LDA", True):  "DALI Licht Ein / Heller",
    ("LDA", False): "DALI Licht Aus / Dunkler",
    ("LC",  True):  "Farblicht Ein/Aus",
    ("LC",  False): "Farblicht Dim",
    ("LCT", True):  "Tunable White Ein/Aus",
    ("LCT", False): "Tunable White Dim",
    ("LCW", True):  "Farblicht RGBW Ein/Aus",
    ("LCW", False): "Farblicht RGBW Dim",
    ("DMX", True):  "DMX Szene weiter",
    ("DMX", False): "DMX Szene zurück",
    ("J",   True):  "Jalousie Auf",
    ("J",   False): "Jalousie Ab",
    ("R",   True):  "Rollladen Auf",
    ("R",   False): "Rollladen Ab",
    ("M",   True):  "Markise Auf",
    ("M",   False): "Markise Ab",
    ("T",   True):  "Vorhang Auf",
    ("T",   False): "Vorhang Ab",
    ("DF",  True):  "Dachfenster Öffnen",
    ("DF",  False): "Dachfenster Schliessen",
    ("F",   True):  "Fenster Öffnen",
    ("F",   False): "Fenster Schliessen",
    ("FG",  True):  "Fliegengitter Auf",
    ("FG",  False): "Fliegengitter Ab",
    ("H",   True):  "Heizung Sollwert +",
    ("H",   False): "Heizung Sollwert –",
    ("S",   True):  "Steckdose Ein/Aus",
    ("S",   False): "Steckdose Ein/Aus",
    ("SD",  True):  "Steckdose Ein / Heller",
    ("SD",  False): "Steckdose Aus / Dunkler",
    ("G",   True):  "Garagentor",
    ("G",   False): "Garagentor",
    ("LW",  True):  "Leinwand Hoch",
    ("LW",  False): "Leinwand Runter",
    ("BL",  True):  "Beamer-Lift Ein",
    ("BL",  False): "Beamer-Lift Aus",
    ("BW",  True):  "Bewässerung Ein/Aus",
    ("BW",  False): "Bewässerung Ein/Aus",
    ("V",   True):  "Ventilator Ein/Aus",
    ("V",   False): "Ventilator Ein/Aus",
    ("LU",  True):  "Lüftung mehr",
    ("LU",  False): "Lüftung weniger",
    ("KL",  True):  "Klimagerät Ein/Aus",
    ("KL",  False): "Klimagerät Temperatur",
    ("MM",  True):  "Multimedia Ein/Aus",
    ("MM",  False): "Multimedia laut/leise",
    ("GS",  True):  "Gong / Sonnerie",
    ("GS",  False): "Gong / Sonnerie",
    ("TE",  True):  "Türe öffnen",
    ("TE",  False): "Türe öffnen",
    ("P",   True):  "Pumpe Ein/Aus",
    ("P",   False): "Pumpe Ein/Aus",
    ("TVL", True):  "TV-Lift Ein",
    ("TVL", False): "TV-Lift Aus",
    ("U",   True):  "Uhr stellen",
    ("U",   False): "Uhr stellen",
    ("GS",  True):  "Gong / Sonnerie",
    ("GS",  False): "Gong / Sonnerie",
    ("EV",  True):  "E-Auto laden",
    ("EV",  False): "E-Auto laden",
}

# Dropdown-Auswahl für leere Taster-Slots – als Excel-Inline-Liste (max. 255 Zeichen)
# Excel erlaubt keine DataValidation-Referenz auf versteckte Blätter;
# daher werden die Optionen direkt als Literal-String in die Formel eingebettet.
_DROPDOWN_OPTIONS = [
    "Licht Ein/Aus", "Licht heller", "Licht dunkler",
    "Jalousie auf", "Jalousie ab",
    "Rollladen auf", "Rollladen ab",
    "Markise auf", "Markise ab",
    "Vorhang auf", "Vorhang ab",
    "Heizung +", "Heizung -",
    "Szene",
    "Steckdose",
    "Lueftung +", "Lueftung -",
    "Klimaanlage", "Garagentor",
    "DALI Licht", "Farblicht",
    "Sonstiges",
]
# Fertige formula1-Zeichenkette (inkl. umschliessender Anführungszeichen)
_DROPDOWN_FORMULA1 = '"' + ",".join(_DROPDOWN_OPTIONS) + '"'


class BauherrFormService:
    """Erzeugt und liest Funktionsdefinitions-Formulare für Bauherren."""

    def __init__(self, project: KnxProject):
        self.project = project

    def _gewerk_label(self, code: str) -> str:
        """Gibt 'Code – Name' zurück, z.B. 'LD – Licht dimmbar'."""
        gewerk = self.project.gewerk_catalog.get(code)
        name = gewerk.name if gewerk else ""
        return f"{code} – {name}" if name else code

    def _room_by_id(self, room_id: str) -> Room | None:
        """Gibt den Raum mit der angegebenen ID zurück."""
        for room in self.project.all_rooms:
            if room.id == room_id:
                return room
        return None

    def _device_product_name(self, be: Bedienelement) -> str:
        """Produktname für die Anzeige: bevorzugt den Live-Wert aus dem per
        physikalischer Adresse verknüpften Device der Topologie.

        be.product_name wird nur einmalig beim Anlegen des Bedienelements aus
        dem Device kopiert (siehe _create_bedienelemente_from_topology) und
        bleibt bei einer späteren Produktzuweisung über die Materialliste
        (die nur ins Device zurückschreibt) sonst veraltet.
        """
        if be.participant_number:
            for area in self.project.topology.areas:
                for line in area.lines:
                    for device in line.devices:
                        if device.physical_address == be.participant_number:
                            if device.product_name:
                                return device.product_name
                            break
        return be.product_name or be.element_type

    def _label_from_ga(self, ga_str: str) -> str:
        """
        Extrahiert den menschlichen Namen aus einer GA-Designation.
        Format: '1/3/5 Raumname Funktionsname' → 'Raumname Funktionsname'
        """
        parts = ga_str.strip().split()
        if parts and "/" in parts[0]:
            return " ".join(parts[1:]) if len(parts) > 1 else ga_str
        return ga_str

    def _human_label(self, sf: SensorFunktion, slot_idx: int) -> str:
        """
        Gibt einen bauherren-lesbaren Funktionsnamen zurück.
        slot_idx: Index in be.funktionen (gerade=oben/primär, ungerade=unten/sekundär).
        """
        # Explizites Label hat Vorrang
        if sf.label:
            return sf.label

        # Direkte GA: beschreibenden Teil extrahieren
        if sf.ga_designation:
            return self._label_from_ga(sf.ga_designation)

        if sf.gewerk_code:
            is_primary = (slot_idx % 2 == 0)
            label = _GEWERK_BTN_LABEL.get((sf.gewerk_code, is_primary))
            if label:
                # Fremdraum-Suffix
                if sf.source_room_id:
                    room = self._room_by_id(sf.source_room_id)
                    room_name = room.name if room else "anderer Raum"
                    return f"{label}  →  {room_name}"
                # Mehrere Elemente desselben Typs: Nummer anhängen
                if sf.element_number > 1:
                    return f"{label} ({sf.element_number})"
                return label
            # Fallback: Gewerk-Name aus Katalog
            gewerk = self.project.gewerk_catalog.get(sf.gewerk_code)
            return gewerk.name if gewerk else sf.gewerk_code
        return ""

    def _parse_button_assignments(
        self, be: Bedienelement, n_rockers: int
    ) -> dict[tuple[int, str], str]:
        """
        Ordnet FunctionAssignments den Button-Slots zu und gibt
        bauherren-lesbare Labels zurück.

        Rückgabe: {(rocker_nr, 'oben'|'unten'): label-text}
        """
        import re
        result: dict[tuple[int, str], str] = {}

        for idx, fa in enumerate(be.function_assignments):
            ch = fa.button_channel.lower()
            # Menschlichen Funktionstext ableiten
            raw = fa.function_ga or fa.description or ""
            parts = raw.split()
            fn = self._label_from_ga(raw) if (parts and "/" in parts[0]) else raw

            m = re.search(r"(\d+)\s*(oben|unten|up|down|auf|ab)", ch)
            if m:
                nr = int(m.group(1))
                direction = "oben" if m.group(2) in ("oben", "up", "auf") else "unten"
                result[(nr, direction)] = fn
                continue

            m2 = re.search(r"t(\d+)", ch)
            if m2:
                nr = int(m2.group(1))
                direction = "oben" if "oben" in ch or "up" in ch else "unten"
                result[(nr, direction)] = fn
                continue

            # Fallback: Index-basiert
            rocker = (idx // 2) + 1
            direction = "oben" if idx % 2 == 0 else "unten"
            result[(rocker, direction)] = fn

        # Wenn keine function_assignments: SensorFunktionen als Fallback (FA-1410)
        if not result and be.funktionen:
            for idx, sf in enumerate(be.funktionen):
                rocker = (idx // 2) + 1
                direction = "oben" if idx % 2 == 0 else "unten"
                result[(rocker, direction)] = self._human_label(sf, idx)

        return result

    def _fill_for(self, fn_text: str, sf: SensorFunktion | None) -> str:
        """Wählt die Hintergrundfarbe je nach Funktionstyp."""
        if not fn_text:
            return _BTN_FILL_WISH       # leer → Wunsch-Slot (Dropdown)
        if sf and sf.ga_designation:
            return _BTN_FILL_GA         # direkte GA → violett
        if sf and sf.label and not sf.gewerk_code:
            # Freitext-Wunsch ohne technische GA-Zuordnung (noch nicht über die
            # Verknüpfungsmatrix aufgelöst) -- wie ein leerer Wunsch behandeln,
            # nicht wie eine fertige Zuweisung.
            return _BTN_FILL_WISH
        if sf and sf.source_room_id:
            return _BTN_FILL_FOREIGN    # Fremdraum → blau
        return _BTN_FILL_ASSIGNED       # eigener Raum → grün

    def _button_label(self, sf: SensorFunktion) -> str:
        """
        Bauherren-lesbarer Label für einen einzelnen Taster-Slot.
        Kein Richtungsbezug (kein 'heller'/'dunkler', 'auf'/'ab') –
        nur der Gewerk-Name und ggf. der Fremdraum.
        """
        if sf.label:
            return sf.label
        if sf.ga_designation:
            return self._label_from_ga(sf.ga_designation)
        if sf.gewerk_code:
            gewerk = self.project.gewerk_catalog.get(sf.gewerk_code)
            label = gewerk.name if gewerk else sf.gewerk_code
            if sf.source_room_id:
                room = self._room_by_id(sf.source_room_id)
                room_name = room.name if room else "anderer Raum"
                return f"{label}  →  {room_name}"
            if sf.element_number > 1:
                return f"{label} ({sf.element_number})"
            return label
        return ""

    # ── Grafische Taster-Darstellung (Einzel-Taster) ──────────────────────────

    def _draw_taster_graphic(self, ws, row: int, col: int,
                             be: Bedienelement,
                             empty_refs: list[str] | None = None) -> int:
        """
        Zeichnet ein Bedienelement als Raster einzelner Taster-Buttons.
        Jeder Slot erscheint als eigenständige Zelle – keine Wippen, keine Pfeile.

        Layout: 2 Spalten, Slots sequenziell von links nach rechts, oben nach unten.
        Koordinaten leerer Slots werden an empty_refs angehängt (für DataValidation).
        """
        import math

        n_buttons = be.channels
        # Alle definierten SensorFunktionen anzeigen, mindestens n_buttons Slots
        n_slots = max(n_buttons, len(be.funktionen))
        n_rows  = math.ceil(n_slots / 2)

        # ── Style-Helfer ───────────────────────────────────────────────────
        def fill(hex_color: str) -> "PatternFill":
            return PatternFill(start_color=hex_color, end_color=hex_color,
                               fill_type="solid")

        thin  = Side(style="thin",   color="BDBDBD")
        outer = Side(style="medium", color="263238")

        NCOLS      = 2
        COL_W      = 20
        ROW_H      = 30
        device_end = col + NCOLS - 1

        for i in range(NCOLS):
            ws.column_dimensions[get_column_letter(col + i)].width = COL_W

        # ── Geräte-Header ─────────────────────────────────────────────────
        pn_str   = be.participant_number or "–"
        dev_name = self._device_product_name(be)
        status   = "Auto" if be.is_auto else "Manuell"
        te_idx   = getattr(be, "taster_index", 1)
        te_prefix = f"Tastereinheit {te_idx}  –  " if be.element_type == "Tastereinheit" else ""

        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=device_end)
        hdr = ws.cell(row=row, column=col,
                      value=f"{te_prefix}{dev_name}   [{pn_str}]   {status}")
        hdr.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        hdr.fill      = fill(_BTN_FILL_HEADER)
        hdr.alignment = Alignment(horizontal="center", vertical="center")
        hdr.border    = Border(left=outer, right=outer, top=outer, bottom=outer)
        ws.row_dimensions[row].height = 18
        row += 1

        # ── Taster-Raster ──────────────────────────────────────────────────
        for slot_idx in range(n_slots):
            grid_row = row + (slot_idx // 2)
            grid_col = col + (slot_idx  % 2)

            is_top_grid_row = (slot_idx // 2 == 0)
            is_bot_grid_row = (slot_idx // 2 == n_rows - 1)
            is_left_col     = (slot_idx  % 2 == 0)
            is_right_col    = (slot_idx  % 2 == 1)

            # Einzel-Slot rechts wenn nur 1 Slot total
            if n_slots == 1:
                is_right_col = True

            ws.row_dimensions[grid_row].height = ROW_H

            sf       = be.funktionen[slot_idx] if slot_idx < len(be.funktionen) else None
            fn_text  = self._button_label(sf) if sf else ""
            bg_color = self._fill_for(fn_text, sf)

            # Taster-Nummer (T1, T2, …) als kleines Präfix
            t_num = f"T{slot_idx + 1}"

            if fn_text:
                cell_val  = f"{t_num}  {fn_text}"
                cell_font = Font(name="Arial", bold=True, size=9,
                                 color=_BTN_FONT_FN)
            else:
                cell_val  = f"{t_num}  → Ihr Wunsch"
                cell_font = Font(name="Arial", bold=False, size=9,
                                 italic=True, color=_BTN_FONT_WISH)

            cell = ws.cell(row=grid_row, column=grid_col, value=cell_val)
            cell.font      = cell_font
            cell.fill      = fill(bg_color)
            cell.alignment = Alignment(horizontal="left", vertical="center",
                                       wrap_text=True, indent=1)
            cell.border    = Border(
                left  = outer if is_left_col  else thin,
                right = outer if is_right_col else thin,
                top   = outer if is_top_grid_row else thin,
                bottom= outer if is_bot_grid_row else thin,
            )

            if not fn_text and empty_refs is not None:
                empty_refs.append(cell.coordinate)

        row += n_rows

        # ── Anmerkungszeile ───────────────────────────────────────────────
        row = self._draw_annotation_row(ws, row, col, device_end)
        return row

    def _draw_annotation_row(self, ws, row: int, col: int,
                              end_col: int) -> int:
        """Zeichnet eine editierbare Anmerkungszeile nach einem Gerät."""
        medium = Side(style="medium", color="263238")
        dotted = Side(style="dotted", color="BDBDBD")

        # Label-Zelle
        lbl = ws.cell(row=row, column=col, value="Anmerkung Bauherr:")
        lbl.font = Font(name="Arial", size=8, italic=True, color="546E7A")
        lbl.fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA",
                               fill_type="solid")
        lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lbl.border = Border(left=medium, top=dotted, bottom=medium)

        # Editierbare Zelle (rest des Device-Blocks, nur mergen wenn >1 Spalte übrig)
        if end_col > col + 1:
            ws.merge_cells(start_row=row, start_column=col + 1,
                           end_row=row, end_column=end_col)
        edit = ws.cell(row=row, column=col + 1, value="")
        edit.fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7",
                                fill_type="solid")
        edit.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        edit.border = Border(right=medium, top=dotted, bottom=medium, left=dotted)

        ws.row_dimensions[row].height = 18
        return row + 2  # 1 Leerzeile danach

    def _draw_freitext_block(self, ws, row: int, col: int,
                             end_col: int) -> int:
        """Zeichnet einen freien Anmerkungsblock am Ende eines Raum-Blatts."""
        medium = Side(style="medium", color="263238")
        dotted = Side(style="dotted", color="BDBDBD")
        thin   = Side(style="thin",   color="BDBDBD")

        # Titel
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=end_col)
        title = ws.cell(row=row, column=col,
                        value="Ihre Anmerkungen zum Raum:")
        title.font = Font(name="Arial", size=10, bold=True, color="1A5276")
        title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 18
        row += 1

        # 4 editierbare Zeilen
        for i in range(4):
            ws.merge_cells(start_row=row + i, start_column=col,
                           end_row=row + i, end_column=end_col)
            edit = ws.cell(row=row + i, column=col, value="")
            edit.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9",
                                    fill_type="solid")
            is_last = (i == 3)
            edit.border = Border(
                left=medium, right=medium,
                top=thin,
                bottom=medium if is_last else dotted,
            )
            edit.alignment = Alignment(horizontal="left", vertical="center",
                                       indent=1)
            ws.row_dimensions[row + i].height = 22

        return row + 4 + 2

    def _draw_legend(self, ws, row: int, col: int) -> int:
        """Zeichnet eine kompakte Farblegende am Ende des Blatts."""
        def fill(hex_color):
            return PatternFill(start_color=hex_color, end_color=hex_color,
                               fill_type="solid")
        entries = [
            (_BTN_FILL_ASSIGNED, "Funktion (eigener Raum)"),
            (_BTN_FILL_FOREIGN,  "Funktion (anderer Raum)"),
            (_BTN_FILL_GA,       "Direkter Sonderwunsch"),
            (_BTN_FILL_WISH,     "Bitte Funktion wählen ↓"),
            (_BTN_FILL_FREE,     "Taste ohne Funktion"),
        ]
        ws.cell(row=row, column=col,
                value="Legende:").font = Font(name="Arial", bold=True, size=8)
        for i, (color, label) in enumerate(entries):
            c = col + 1 + i * 2
            sample = ws.cell(row=row, column=c, value="  ")
            sample.fill = fill(color)
            ws.cell(row=row, column=c + 1, value=label).font = Font(
                name="Arial", size=8, color="555555"
            )
        return row + 2

    # ── Formular generieren ────────────────────────────────────────────────────

    def generate_form(self, filepath: str):
        """Erzeugt ein Excel-Formular pro Raum für den Bauherren (FA-1501)."""
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl wird für Excel-Export benoetigt.")

        excel = ExcelGenerator(
            title="Funktionsdefinition",
            company=self.project.project_info.company_name
                    if hasattr(self.project.project_info, "company_name")
                    else "",
            project_name=self.project.name,
        )

        # ── Übersichtsblatt ────────────────────────────────────────────────
        excel.add_header()
        excel.add_heading("Tastenbelegung – Bauherren-Formular", level=1)
        excel.add_empty_row()
        excel.add_paragraph(
            "Dieses Formular zeigt die geplante Tastenbelegung pro Raum.\n"
            "Grün = Funktion zugewiesen  |  Blau = Funktion aus anderem Raum  "
            "|  Violett = Direkter Sonderwunsch  |  Gelb = Bitte Funktion wählen\n"
            "Tippen Sie in einen gelben Slot – eine Auswahlliste erscheint."
        )
        excel.add_empty_row()

        rooms = self.project.all_rooms
        headers = ["Raum-Nr.", "Raumname", "Bedienelemente", "Gewerke"]
        rows = []
        for room in rooms:
            gewerke = ", ".join(
                self._gewerk_label(ga.gewerk_code)
                for ga in room.gewerk_assignments
            )
            rows.append([
                room.number,
                room.name,
                str(len(room.bedienelemente)),
                gewerke or "–",
            ])
        excel.add_table(headers, rows, col_widths=[12, 25, 16, 60])

        # ── Pro Raum ein eigenes Blatt ─────────────────────────────────────
        for room in rooms:
            if not room.bedienelemente:
                continue

            raw = f"{room.number} {room.name}"
            for ch in r'\/?*[]':
                raw = raw.replace(ch, "-")
            raw = raw.replace(":", "-")
            sheet_name = raw[:31]
            excel.add_sheet(sheet_name)
            ws = excel._current_sheet

            # Raumtitel
            from openpyxl.styles import Font as OFont, PatternFill as OFill, Alignment as OAlign
            ws.merge_cells("A1:L1")
            title_cell = ws["A1"]
            title_cell.value = f"Raum {room.number}:  {room.name}"
            title_cell.font = OFont(name="Arial", bold=True, size=14,
                                    color="FFFFFF")
            title_cell.fill = OFill(start_color="3A6B19", end_color="3A6B19",
                                    fill_type="solid")
            title_cell.alignment = OAlign(horizontal="left", vertical="center",
                                          indent=1)
            ws.row_dimensions[1].height = 24

            # Projekt-Info
            ws.merge_cells("A2:L2")
            info = ws["A2"]
            info.value = (
                f"Projekt: {self.project.name}   |   "
                f"Datum: {datetime.now().strftime('%d.%m.%Y')}   |   "
                f"KNX Arranger"
            )
            info.font = OFont(name="Arial", size=9, color="555555", italic=True)
            ws.row_dimensions[2].height = 14
            current_row = 4

            # Gewerke-Liste
            if room.gewerk_assignments:
                ws.cell(row=current_row, column=1,
                        value="Installierte Gewerke:").font = OFont(
                    name="Arial", bold=True, size=10, color="1A5276"
                )
                current_row += 1
                for ga in room.gewerk_assignments:
                    gewerk = self.project.gewerk_catalog.get(ga.gewerk_code)
                    name = gewerk.name if gewerk else ""
                    ws.cell(row=current_row, column=1,
                            value=f"  {ga.gewerk_code}  –  {name}  ×{ga.count}"
                            ).font = OFont(name="Arial", size=9)
                    current_row += 1
                current_row += 1

            ws.cell(row=current_row, column=1,
                    value="Tastenbelegung:").font = OFont(
                name="Arial", bold=True, size=10, color="1A5276"
            )
            current_row += 1

            # Hinweis für den Bauherrn
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=4,
            )
            hint = ws.cell(
                row=current_row, column=1,
                value=(
                    "Gelbe Felder: Klicken Sie in das Feld und wählen Sie "
                    "Ihre gewünschte Funktion aus der Liste."
                )
            )
            hint.font = OFont(name="Arial", size=9, italic=True, color="E65100")
            hint.fill = OFill(start_color="FFF8E1", end_color="FFF8E1",
                              fill_type="solid")
            ws.row_dimensions[current_row].height = 14
            current_row += 2

            legend_drawn = False
            device_end_col = 2  # col=1, COLS_PER_PAIR=2 → end=2

            # Leere Slot-Koordinaten sammeln (für DataValidation nach dem Zeichnen)
            empty_refs: list[str] = []

            for be in room.bedienelemente:
                current_row = self._draw_taster_graphic(
                    ws, current_row, col=1, be=be, empty_refs=empty_refs
                )
                if not legend_drawn:
                    current_row = self._draw_legend(ws, current_row, col=1)
                    legend_drawn = True

            # DataValidation nur anlegen wenn mindestens ein leerer Slot existiert.
            # Ein leeres sqref wäre ungültiges OOXML und führt zu XML-Fehlern in Excel.
            if empty_refs:
                from openpyxl.worksheet.datavalidation import DataValidation
                dv = DataValidation(
                    type="list",
                    formula1=_DROPDOWN_FORMULA1,
                    showDropDown=False,
                    allow_blank=True,
                    showErrorMessage=False,
                    showInputMessage=True,
                    promptTitle="Wunschfunktion",
                    prompt="Bitte Funktion aus Liste waehlen.",
                )
                dv.sqref = " ".join(empty_refs)
                ws.add_data_validation(dv)

            # Freier Anmerkungsblock am Ende jedes Raum-Blatts
            current_row = self._draw_freitext_block(
                ws, current_row, col=1, end_col=device_end_col
            )

        excel.save(filepath)
        logger.info(f"Bauherr-Formular erstellt: {filepath}")

    def import_form(self, filepath: str) -> int:
        """
        Liest ein ausgefülltes Bauherr-Formular zurück (FA-1503).

        Parsing-Strategie für das Einzel-Taster-Format:
        - Geräte-Header: Zelle mit '[…]'-Muster → nächstes Gerät
        - Taster-Zellen: Inhalt beginnt mit 'T<N>  ' (z.B. 'T3  Jalousie')
          → sf_idx = N-1, Wert = Text nach dem Prefix
        - '→ Ihr Wunsch' bedeutet nicht ausgefüllt → überspringen
        - Neue Wahl → SensorFunktion.label aktualisieren oder neue SF anlegen
        """
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl wird für Excel-Import benoetigt.")

        import re
        from openpyxl import load_workbook
        from ..models.building import SensorFunktion as SF

        wb = load_workbook(filepath, data_only=True)
        rooms_by_number: dict[str, Room] = {
            r.number: r for r in self.project.all_rooms
        }
        imported_count = 0
        _HDR_SKIP = ("anmerkung", "legende", "raum", "projekt")
        _T_RE = re.compile(r'^T(\d+)\s{1,3}(.*)$')

        def _apply(be: Bedienelement, sf_idx: int, val: str):
            nonlocal imported_count
            if not val or "Ihr Wunsch" in val:
                return
            if sf_idx < len(be.funktionen):
                sf = be.funktionen[sf_idx]
                auto = self._button_label(sf)
                if val != auto:
                    sf.label = val
                    imported_count += 1
                    be.is_auto = False
            else:
                while len(be.funktionen) < sf_idx:
                    be.funktionen.append(SF())
                # NUR das Freitext-Label setzen -- ga_designation bleibt leer,
                # bis ein Planer dem Wunsch ueber die Verknuepfungsmatrix
                # (FA-2503) eine echte Gruppenadresse zuweist. ga_designation
                # mit dem rohen Bauherr-Freitext zu befuellen wuerde
                # SensorService._expand_funktionen dazu bringen, ihn wie eine
                # echte GA zu behandeln (sf.ga_designation ist der alleinige
                # Trigger fuer die "Direkte GA"-Variante).
                be.funktionen.append(SF(label=val))
                imported_count += 1
                # Wunsch markiert das Bedienelement als manuell konfiguriert, damit
                # auto_assign_functions die Funktionsliste bei der nachfolgenden
                # Neuberechnung nicht verwirft (FA-1410: is_auto=False wird erhalten).
                be.is_auto = False

        for sheet_name in wb.sheetnames:
            if sheet_name == "Funktionsdefinition"[:31]:
                continue

            ws = wb[sheet_name]
            room_number = (sheet_name.split(" ")[0]
                           if " " in sheet_name else sheet_name)
            room = rooms_by_number.get(room_number)
            if not room:
                continue

            device_idx = -1

            for row_cells in ws.iter_rows(min_row=1, values_only=False):
                for cell in row_cells[:2]:   # nur Spalten A+B
                    v = str(cell.value or "").strip()
                    if not v:
                        continue

                    # Geräte-Header
                    if (re.search(r'\[.+?\]', v)
                            and not any(k in v.lower() for k in _HDR_SKIP)):
                        device_idx += 1
                        break

                    # Taster-Zelle: "T3  Jalousie"
                    m = _T_RE.match(v)
                    if m and 0 <= device_idx < len(room.bedienelemente):
                        sf_idx = int(m.group(1)) - 1
                        fn_val = m.group(2).strip()
                        _apply(room.bedienelemente[device_idx], sf_idx, fn_val)

        wb.close()
        logger.info(f"Bauherr-Formular importiert: {imported_count} Änderungen")
        return imported_count


def _get_button_count(sensor_type: str) -> int:
    """Ermittelt die Tastenanzahl aus dem Sensortyp."""
    sensor_lower = sensor_type.lower()
    if "8-fach" in sensor_lower or "8fach" in sensor_lower:
        return 8
    if "6-fach" in sensor_lower or "6fach" in sensor_lower:
        return 6
    if "4-fach" in sensor_lower or "4fach" in sensor_lower:
        return 4
    if "2-fach" in sensor_lower or "2fach" in sensor_lower:
        return 2
    if "1-fach" in sensor_lower or "1fach" in sensor_lower:
        return 1
    return 4  # Standard
