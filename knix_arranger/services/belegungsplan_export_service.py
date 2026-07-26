"""
ETS-Belegungsplan Excel-Export.

Exportiert den Sensor/Taster-Belegungsplan als formatierte .xlsx-Datei,
die dem ETS-Programmierer als Arbeitsvorlage dient.
"""
from __future__ import annotations
import logging
from datetime import datetime

from .belegungsplan_service import _split_button_channel

logger = logging.getLogger("knix_arranger.belegungsplan_export")

# Spalten Tab "Taster & Sensoren"  (11 Spalten A–K)
_SENSOR_HEADERS = [
    "Stockwerk",        # A
    "Zone",             # B
    "Raum-Nr.",         # C
    "Raum",             # D
    "Gerätetyp",        # E
    "Phys. Adresse",    # F  ← Monospace
    "Taste",            # G  (kombiniert: "Taste 1", "Status", "Sollwert" …)
    "Aktion",           # H  ("kurz", "lang", "Rückmeld.", …)
    "Funktion",         # I
    "GA-Bezeichnung",   # J
    "GA-Adresse",       # K  ← Monospace
    "DPT",              # L  ← Monospace
]
_SENSOR_WIDTHS    = [18, 20, 10, 22, 18, 14, 20, 12, 28, 32, 12, 14]
_SENSOR_LAST_COL  = "L"   # für Merge-Ranges und AutoFilter
_SENSOR_MONO_COLS = (6, 11, 12)   # F, K, L
_SENSOR_WARN_COLS = (10, 11, 12)  # J, K, L  (warn wenn kein GA)

# Spalten Tab "Aktoren"  (13 Spalten A–M)
_ACTOR_HEADERS = [
    "Linie",                # A
    "UV-Standort",          # B
    "Gerätetyp (Aktor)",    # C
    "Phys. Adresse",        # D  ← Monospace
    "Kanal",                # E
    "Raum-Nr.",             # F
    "Zone",                 # G
    "Raum",                 # H
    "Gewerk",               # I
    "Funktion",             # J
    "GA-Bezeichnung",       # K
    "GA-Adresse",           # L  ← Monospace
    "DPT",                  # M  ← Monospace
]
_ACTOR_WIDTHS    = [20, 18, 22, 14, 8, 10, 20, 22, 10, 12, 32, 12, 14]
_ACTOR_LAST_COL  = "M"   # für Merge-Ranges und AutoFilter
_ACTOR_MONO_COLS = (4, 12, 13)    # D, L, M

# Farben
_COLOR_HEADER_BG = "1565C0"   # KNX-Blau
_COLOR_HEADER_FG = "FFFFFF"
_COLOR_ROW_ALT   = "F5F5F5"  # hellgrau für jede 2. Zeile
_COLOR_ROW_NORM  = "FFFFFF"
_COLOR_GA_EMPTY  = "FFCDD2"  # hellrot wenn GA-Adresse fehlt


class BelegungsplanExportService:
    """Exportiert BelegungsplanData als .xlsx-Datei."""

    def export_xlsx(self, data, filepath: str) -> None:
        """
        Schreibt den Belegungsplan als Excel-Datei.

        Args:
            data:     BelegungsplanData aus BelegungsplanService.generate()
            filepath: Zielpfad (inkl. .xlsx-Erweiterung)
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise RuntimeError("openpyxl nicht verfügbar – bitte installieren.") from e

        wb = openpyxl.Workbook()

        # ── Tab "Taster & Sensoren" ──
        ws = wb.active
        ws.title = "Taster & Sensoren"

        # Titel-Zeile
        ws.merge_cells(f"A1:{_SENSOR_LAST_COL}1")
        title_cell = ws["A1"]
        title_cell.value = f"ETS-Belegungsplan: {data.project_name}"
        title_cell.font = Font(name="Calibri", size=13, bold=True, color=_COLOR_HEADER_FG)
        title_cell.fill = PatternFill("solid", fgColor=_COLOR_HEADER_BG)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        # Datum-Zeile
        ws.merge_cells(f"A2:{_SENSOR_LAST_COL}2")
        date_cell = ws["A2"]
        date_cell.value = f"Erstellt: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        date_cell.font = Font(name="Calibri", size=9, italic=True, color="666666")
        date_cell.alignment = Alignment(horizontal="right")

        # Header-Zeile (Zeile 3)
        header_fill = PatternFill("solid", fgColor=_COLOR_HEADER_BG)
        header_font = Font(name="Calibri", size=10, bold=True, color=_COLOR_HEADER_FG)
        thin_border = Border(
            bottom=Side(style="thin", color="AAAAAA"),
            right=Side(style="thin", color="AAAAAA"),
        )
        for col, header in enumerate(_SENSOR_HEADERS, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[3].height = 18

        # AutoFilter auf Header-Zeile + Zeilen einfrieren
        ws.auto_filter.ref = f"A3:{_SENSOR_LAST_COL}3"
        ws.freeze_panes = "A4"

        # Datenzeilen
        norm_fill = PatternFill("solid", fgColor=_COLOR_ROW_NORM)
        alt_fill  = PatternFill("solid", fgColor=_COLOR_ROW_ALT)
        warn_fill = PatternFill("solid", fgColor=_COLOR_GA_EMPTY)
        data_font = Font(name="Calibri", size=10)
        addr_font = Font(name="Courier New", size=10)

        prev_room = None
        for i, row in enumerate(data.sensor_rows):
            excel_row = i + 4
            is_alt = (i % 2 == 1)
            row_fill = alt_fill if is_alt else norm_fill
            no_ga = not row.ga_address

            action_label = "Rückmeld." if row.is_feedback else row.action_type
            values = [
                row.floor_name,
                row.zone_name,
                row.room_number,
                row.room_name,
                row.sensor_type,
                row.physical_address,
                row.taste_label,
                action_label,
                row.function,
                row.ga_designation,
                row.ga_address,
                row.dpt,
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=excel_row, column=col, value=val)
                cell.font = addr_font if col in _SENSOR_MONO_COLS else data_font
                cell.fill = warn_fill if (no_ga and col in _SENSOR_WARN_COLS) else row_fill
                cell.alignment = Alignment(vertical="center")
                cell.border = Border(
                    bottom=Side(style="hair", color="DDDDDD"),
                    right=Side(style="hair", color="DDDDDD"),
                )

            # Trennlinie bei Raumwechsel
            if prev_room and prev_room != row.room_number:
                for col in range(1, len(_SENSOR_HEADERS) + 1):
                    ws.cell(row=excel_row, column=col).border = Border(
                        top=Side(style="thin", color="AAAAAA"),
                        bottom=Side(style="hair", color="DDDDDD"),
                        right=Side(style="hair", color="DDDDDD"),
                    )
            prev_room = row.room_number

        # Spaltenbreiten
        for col, width in enumerate(_SENSOR_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Legende am Ende
        legend_row = len(data.sensor_rows) + 6
        ws.cell(row=legend_row, column=1, value="Legende:").font = Font(bold=True, size=9)
        ws.cell(row=legend_row + 1, column=1,
                value="Rot hinterlegt = GA-Adresse nicht gefunden (GA nicht generiert oder Name abweichend)")
        ws.cell(row=legend_row + 1, column=1).font = Font(size=9, color="C62828")

        # Hinweis
        hint_row = legend_row + 3
        ws.merge_cells(f"A{hint_row}:{_SENSOR_LAST_COL}{hint_row}")
        ws.cell(row=hint_row, column=1,
                value="Hinweis: Diese Tabelle zeigt die empfohlenen GA-Zuordnungen für die ETS-Programmierung. "
                      "Status-/Rückmelde-Objekte (RM) müssen als lesende KOs konfiguriert werden.").font = Font(
            size=9, italic=True, color="555555")

        # ── Tab "Aktoren" ──
        if data.actor_rows:
            ws2 = wb.create_sheet("Aktoren")

            # Titel
            ws2.merge_cells(f"A1:{_ACTOR_LAST_COL}1")
            t2 = ws2["A1"]
            t2.value = f"ETS-Belegungsplan Aktoren: {data.project_name}"
            t2.font = Font(name="Calibri", size=13, bold=True, color=_COLOR_HEADER_FG)
            t2.fill = PatternFill("solid", fgColor=_COLOR_HEADER_BG)
            t2.alignment = Alignment(horizontal="center", vertical="center")
            ws2.row_dimensions[1].height = 22

            # Datum
            ws2.merge_cells(f"A2:{_ACTOR_LAST_COL}2")
            d2 = ws2["A2"]
            d2.value = f"Erstellt: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            d2.font = Font(name="Calibri", size=9, italic=True, color="666666")
            d2.alignment = Alignment(horizontal="right")

            # Header
            for col, header in enumerate(_ACTOR_HEADERS, start=1):
                cell = ws2.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
            ws2.row_dimensions[3].height = 18
            ws2.auto_filter.ref = f"A3:{_ACTOR_LAST_COL}3"
            ws2.freeze_panes = "A4"

            # Datenzeilen
            prev_actor = None
            for i, row in enumerate(data.actor_rows):
                excel_row = i + 4
                is_alt = (i % 2 == 1)
                row_fill = alt_fill if is_alt else norm_fill

                values = [
                    f"{row.area_number}.{row.line_number} {row.line_name}",
                    row.uv_location,
                    row.actor_type,
                    row.physical_address,
                    row.channel_number,
                    row.room_number,
                    row.zone_name,
                    row.room_name,
                    row.gewerk_code,
                    row.function_name,
                    row.ga_designation,
                    row.ga_address,
                    row.dpt,
                ]
                for col, val in enumerate(values, start=1):
                    cell = ws2.cell(row=excel_row, column=col, value=val)
                    cell.font = addr_font if col in _ACTOR_MONO_COLS else data_font
                    cell.fill = row_fill
                    cell.alignment = Alignment(vertical="center")
                    cell.border = Border(
                        bottom=Side(style="hair", color="DDDDDD"),
                        right=Side(style="hair", color="DDDDDD"),
                    )

                # Trennlinie bei Aktor-Wechsel
                actor_key = row.physical_address
                if prev_actor and prev_actor != actor_key:
                    for col in range(1, len(_ACTOR_HEADERS) + 1):
                        ws2.cell(row=excel_row, column=col).border = Border(
                            top=Side(style="thin", color="AAAAAA"),
                            bottom=Side(style="hair", color="DDDDDD"),
                            right=Side(style="hair", color="DDDDDD"),
                        )
                prev_actor = actor_key

            # Spaltenbreiten
            for col, width in enumerate(_ACTOR_WIDTHS, start=1):
                ws2.column_dimensions[get_column_letter(col)].width = width

            # Hinweis
            hint2 = len(data.actor_rows) + 6
            ws2.merge_cells(f"A{hint2}:{_ACTOR_LAST_COL}{hint2}")
            ws2.cell(row=hint2, column=1,
                     value="Hinweis: Spalte 'Kanal' zeigt den Ausgangskanal des Aktors (1, 2, 3 …). "
                           "Der Elektriker schliesst das jeweilige Gewerk an diesen Klemmenanschluss an."
                     ).font = Font(size=9, italic=True, color="555555")

        wb.save(filepath)
        logger.info(
            f"Belegungsplan exportiert: {filepath} "
            f"({len(data.sensor_rows)} Sensor-Zeilen, {len(data.actor_rows)} Aktor-Zeilen)"
        )

    def export_pdf(self, data, filepath: str, company_profile=None, project_info=None) -> None:
        """
        FA-2505: Exportiert den Belegungsplan als PDF.

        Args:
            data:            BelegungsplanData aus BelegungsplanService.generate()
            filepath:        Zielpfad (inkl. .pdf-Erweiterung)
            company_profile: Optionales CompanyProfile fuer Header/Footer
            project_info:    Optionales ProjectInfo fuer Header
        """
        from ..utils.pdf_generator import PdfGenerator

        pdf = PdfGenerator(
            title="ETS-Belegungsplan",
            company_profile=company_profile,
            project_info=project_info,
        )

        # ── Sensoren / Taster ──────────────────────────────────────────────────
        pdf.add_heading(f"ETS-Belegungsplan: {data.project_name}", level=1)
        pdf.add_heading("Taster & Sensoren", level=2)

        s_headers = ["Raum-Nr.", "Raumname", "Sensor-Typ", "Taste", "Kn.", "Funktion",
                     "GA-Bezeichnung", "GA-Adresse", "DPT"]
        # 495 pts usable width (595 − 2×50); Breiten an typische Inhaltslängen angepasst
        # (Taste/Funktion brauchten mehr Platz, Kn. deutlich weniger).
        s_widths = [32.0, 58.0, 62.0, 55.0, 18.0, 78.0, 112.0, 34.0, 46.0]
        s_rows = []
        for row in data.sensor_rows:
            taste, kanal = _split_button_channel(row.taste_label)
            s_rows.append([
                row.room_number, row.room_name, row.sensor_type,
                taste, kanal, row.function,
                row.ga_designation, row.ga_address, row.dpt,
            ])
        if s_rows:
            pdf.add_table(s_headers, s_rows, col_widths=s_widths)
        else:
            pdf.add_paragraph("(Keine Sensor/Taster-Zeilen vorhanden)")

        # ── Aktoren ───────────────────────────────────────────────────────────
        if data.actor_rows:
            pdf.add_page_break()
            pdf.add_heading("Aktoren", level=2)

            a_headers = ["Linie", "Aktor-Typ", "Phys.Adr.", "Kn.", "Raum-Nr.",
                         "Gew.", "Funktion", "GA-Bezeichnung", "GA-Adresse", "DPT"]
            # Linie/Aktor-Typ/GA-Bezeichnung sind meist die längsten Inhalte und
            # liefen zuvor über; DPT/Gew./GA-Adresse waren überdimensioniert.
            a_widths = [73.0, 73.0, 38.0, 16.0, 36.0, 20.0, 42.0, 116.0, 36.0, 45.0]
            a_rows = []
            for row in data.actor_rows:
                a_rows.append([
                    f"{row.area_number}.{row.line_number} {row.line_name}",
                    row.actor_type, row.physical_address, row.channel_number,
                    row.room_number, row.gewerk_code, row.function_name,
                    row.ga_designation, row.ga_address, row.dpt,
                ])
            pdf.add_table(a_headers, a_rows, col_widths=a_widths)

        pdf.save(filepath)
        logger.info(
            f"Belegungsplan PDF exportiert: {filepath} "
            f"({len(data.sensor_rows)} Sensor-Zeilen, {len(data.actor_rows)} Aktor-Zeilen)"
        )
