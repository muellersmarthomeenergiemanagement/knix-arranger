"""
ETS6 XLSX Topologie-Report Import (FA-511 bis FA-520)
Liest ETS6-Topologie-Reports mit Zusatzwahl "Objekte" ein.

Spaltenstruktur (0-basiert) der ETS6-XLSX-Datei, Referenzlayout:
  Spalte  6: Schlüssel (Adresse / KO-Nummer / Topologie-ID)
  Spalte 11: Medium (TP/IP) bei Topologie-Zeilen; Hersteller bei Geräten
  Spalte 13: Name bei Topologie-Zeilen (Bereich/Linie/Backbone)
  Spalte 18: Einbauort (Geräte-Subzeile "Raum")
  Spalte 19: Bestellnummer bei Geräten
  Spalte 23: Produkt bei Geräten
  Spalte 37: Applikationsprogramm bei Geräten
  Spalte  9: KO-Name
  Spalte 17: KO-Objektfunktion
  Spalte 25: KO-Prioritaet
  Spalte 30: KO-Flags
  Spalte 33: KO-Datentyp
  Spalte 38: KO verbundene Gruppenadresse(n) / Fortsetzungszeilen

ETS6 verschiebt diese Indizes je nach gewählten Export-Zusatzspalten
(z.B. mehr Metadatenfelder) -- die Deltas zwischen den Spalten sind dabei
nicht einheitlich. Die obigen Werte dienen daher nur noch als Fallback;
die tatsächlichen Spalten werden pro Datei anhand der Kopfzeilen-Texte
ermittelt (siehe `_resolve_topology_columns`). Metadaten (Projekt/Datum)
werden nicht über feste Spalten gelesen, sondern als Label:Wert-Paar
in derselben Zeile gesucht (siehe `_extract_metadata`).
"""
from __future__ import annotations
import logging
import os
import re
from typing import Optional

from ..models.topology import (
    Topology, Area, Line, Device, CommunicationObject,
)
from ..models.building import (
    Areal, Building, Wing, Floor, Apartment, Room,
    STANDARD_FLOOR_NAMES, FLOOR_TO_MAIN_GROUP,
)
from ..models.group_address import (
    GroupAddressStructure, MainGroup, MiddleGroup, GroupAddress,
)

logger = logging.getLogger("knix_arranger.xlsx_import")

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Spaltenindizes Topologie-Report (0-basiert)
COL_KEY = 6          # Adresse / KO-Nummer / Topologie-ID
COL_MEDIUM = 11      # TP / IP  (Topologie) oder Hersteller (Gerät)
COL_TOPO_NAME = 13   # Bereich/Linien-Name
COL_LOCATION = 18    # Einbauort (Geräte-Subzeile)
COL_ORDER_NR = 19    # Bestellnummer Gerät
COL_PRODUCT = 23     # Produkt-Bezeichnung Gerät
COL_APP_PROG = 37    # Applikationsprogramm Gerät
COL_KO_NAME = 9      # KO-Name
COL_KO_FUNC = 17     # KO-Funktion
COL_KO_PRIO = 25     # KO-Prioritaet
COL_KO_FLAGS = 30    # KO-Flags
COL_KO_DT = 33       # KO-Datentyp
COL_GA = 38          # Verbundene Gruppenadresse(n)

# Kopfzeilen-Text -> (Delta Werte-Spalte minus Kopfzeilen-Spalte).
# Kalibriert am Referenzlayout oben; das Delta entsteht durch Merged-Cell-
# Header (die Beschriftung sitzt nicht immer exakt über der Werte-Spalte)
# und ist projektübergreifend stabil, auch wenn ETS6 zusätzliche Spalten
# einfügt und damit alle Indizes nach rechts verschiebt.
_TOPOLOGY_HEADER_ANCHORS: dict[str, tuple[str, int]] = {
    "KEY": ("Adresse", 0),
    "MEDIUM": ("Hersteller", 1),
    "ORDER_NR": ("Bestellnummer", 0),
    "PRODUCT": ("Produkt", -2),
    "APP_PROG": ("Applikation", 0),
    "LOCATION": ("Raum", 3),
    "KO_NAME": ("Name", 0),
    "KO_FUNC": ("Objektfunktion", 0),
    "KO_PRIO": ("Priorität", 1),
    "KO_FLAGS": ("Flags", 0),
    "KO_DT": ("Datentyp", 0),
    "GA": ("Verbunden mit", 0),
}

# Spaltenindizes GA-Report (0-basiert)
COL_GAR_ADDR = 4        # HG-Nr / MG-Adresse / GA-Adresse / Geräteadresse / KO
COL_GAR_NAME = 8        # Name bei HG, MG und GA-Zeilen
COL_GAR_DTYPE = 22      # Datentyp bei GA-Zeilen
COL_GAR_META_LABEL = 7  # Metadaten-Label
COL_GAR_META_VALUE = 18 # Metadaten-Wert
COL_GAR_PRODUCT = 6     # Produktname bei Geräte-Zeilen
COL_GAR_LOCATION = 19   # Einbauort bei Geräte-Zeilen ("04 Schlafen" / "UV2 (Steigzone)")
COL_GAR_KO_DTYPE = 20   # Datentyp bei KO-Zeilen
COL_GAR_KO_PRIO = 23    # Priorität bei KO-Zeilen
COL_GAR_KO_FLAGS = 25   # Flags bei KO-Zeilen
COL_GAR_KO_GAS = 28     # Vollständige verbundene GAs bei KO-Zeilen (alle, nicht nur aktuelle)

# Regex
_PHYS_ADDR_RE = re.compile(r"^(\d{1,2})\.(\d{1,2})\.(\d{1,3})$")
_POWER_SUPPLY_RE = re.compile(r"^(\d{1,2})\.(\d{1,2})\.-$")
_LINE_ADDR_RE = re.compile(r"^(\d{1,2})\.(\d{1,2})$")
_GA_RE = re.compile(r"\d+/\d+/\d+")
# GA-Report Zeilenerkennung
_GAR_HG_RE = re.compile(r"^\d{1,2}$")                     # Hauptgruppe: '0', '1', '12'
_GAR_MG_RE = re.compile(r"^\d{1,2}/\d{1,2}$")            # Mittelgruppe: '0/0', '1/3'
_GAR_GA_RE = re.compile(r"^\d{1,2}/\d{1,2}/\d{1,3}$")    # Gruppenadresse: '0/0/100'
_GAR_KO_RE = re.compile(r"^\d+:")                          # KO-Zeile: '1: Ausgang B', '32: G1,...'
# Einbauort-Raum-Format: "04  Schlafen" → (room_nr, room_name)
_EINBAUORT_ROOM_RE = re.compile(r"^(\d{2})\s+(.+)$")
# GA-Name-Muster für Gebäudestruktur-Ableitung: z.B. "S.OG.04.01_ea ( Beschreibung )"
_GA_FLOOR_ROOM_RE = re.compile(
    r"[A-Z]{1,2}\.([A-Z0-9]{2,5})\.(\d{2})\.\d{2}[_a-z]*\s*\(\s*([^)]{2,40})\)"
)
# Einfacheres Muster für Stockwerk/Raum-Extraktion (ohne Klammerpflicht)
_GA_FLOOR_ROOM_SIMPLE_RE = re.compile(
    r"^[A-Z]{1,4}\.([A-Z0-9]{2,5})\.(\d{2})\.\d{2}"
)
# Worte, die KEIN Raumname sind (technische/direktionale Bezeichnungen)
_NON_ROOM_WORDS = frozenset({
    "spots", "indirekt", "berg", "rotten", "mitte", "wand", "ost", "west",
    "nord", "sued", "sued", "links", "rechts", "oben", "unten", "seite",
    "status", "schalten", "dimmen", "jalousie", "heizung", "lueftung",
    "szene", "nacht", "tag", "abwesend", "anwesend", "abwesenheit",
})

# 3-stellige Raumnummer ohne Stockwerk-Buchstaben (EFH-Konvention
# "{Stockwerk}{nn}", z.B. "L.004.1_ea" -> Stockwerk 0, Raum 04).
# Eine optionale Klammer-Beschreibung wird mitgenommen, falls vorhanden.
# Kein '^'-Anker: wird per finditer() auf Zellentext "Adresse Bezeichnung..."
# angewendet, die Bezeichnung beginnt also nicht am Stringanfang.
_GA_DIGIT_ROOM_RE = re.compile(
    r"\b[A-Z]{1,4}\.(\d)(\d{2})\.\d{1,2}[_a-z]*\s*(?:\(\s*([^)]{2,40})\))?"
)
_GA_DIGIT_ROOM_SIMPLE_RE = re.compile(r"^[A-Z]{1,4}\.(\d)(\d{2})\.")
# Seriennummern im Format "XXXX:XXXXXXXX" (Hex) -- kein GA-Hinweis
_SERIAL_NUMBER_RE = re.compile(r"^[0-9A-Fa-f]{4}:[0-9A-Fa-f]+$")

# Produktname-Schluesselwoerter zur Geraete-Klassifizierung (analog
# knxproj_import_service._infer_device_type), als Ergaenzung zur reinen
# Adress-Heuristik, die bei Tastern/Sensoren mit niedriger Adresse irrt.
_INFRA_KEYWORDS = (
    "remote access", "ip router", "ip interface", "usb interface",
    "knx router", "knx interface", "knx gateway", "knxnet",
    "line coupler", "area coupler", "backbone coupler",
    "power supply", "speisegerät", "netzteil",
)
_SENSOR_KEYWORDS = (
    "sensor", "button", "taster", "push", "presence", "präsenz",
    "temperature", "temperatur", "weather", "wetter", "detector",
    "bewegungsmelder", "raumthermostat", "thermostat",
)
_ACTOR_KEYWORDS = (
    "actuator", "aktor", "switch act", "schaltakt", "dimm",
    "jalousie", "shutter", "blind", "hvac", "heating ctrl",
    "dali", "rgb led", "led driver",
)


def _ga_sort_key(addr: str) -> tuple[int, int, int]:
    """Sortierschluessel für GA-Adressen (H/M/S)."""
    parts = addr.split("/")
    try:
        return (int(parts[0]), int(parts[1]), int(parts[2]))
    except (ValueError, IndexError):
        return (999, 999, 999)


def _cell(row: tuple, idx: int) -> str:
    """Gibt den Zellenwert als bereinigten String zurück."""
    if idx < len(row) and row[idx] is not None:
        return str(row[idx]).strip()
    return ""


class XlsxImportService:
    """Importiert ETS6-Topologie-Reports (XLSX) (FA-511)."""

    def _resolve_topology_columns(self, rows: list) -> dict[str, int]:
        """
        Ermittelt die tatsächlichen Spaltenindizes des Topologie-Reports
        anhand der Kopfzeilen-Texte ("Adresse", "Hersteller", "Produkt", ...).

        ETS6 exportiert je nach gewählten Zusatzspalten unterschiedlich
        breite Reports (z.B. mit mehr Metadatenfeldern) -- feste Indizes
        brechen dann, da die Verschiebung zwischen den Spaltengruppen
        nicht einheitlich ist. Die Kopfzeilen-Texte bleiben stabil; ihr
        Versatz zur Werte-Spalte ist über `_TOPOLOGY_HEADER_ANCHORS`
        kalibriert. Wird ein Header nicht gefunden (z.B. abweichendes
        Exportformat), greift der Referenz-Index als Fallback.
        """
        cols: dict[str, int] = {
            "KEY": COL_KEY, "MEDIUM": COL_MEDIUM, "TOPO_NAME": COL_TOPO_NAME,
            "LOCATION": COL_LOCATION, "ORDER_NR": COL_ORDER_NR,
            "PRODUCT": COL_PRODUCT, "APP_PROG": COL_APP_PROG,
            "KO_NAME": COL_KO_NAME, "KO_FUNC": COL_KO_FUNC,
            "KO_PRIO": COL_KO_PRIO, "KO_FLAGS": COL_KO_FLAGS,
            "KO_DT": COL_KO_DT, "GA": COL_GA,
        }

        header_idx: dict[str, int] = {}
        for row in rows[:80]:
            if not row:
                continue
            for idx, cell in enumerate(row):
                if isinstance(cell, str):
                    text = cell.strip()
                    if text and text not in header_idx:
                        header_idx[text] = idx

        for key, (text, delta) in _TOPOLOGY_HEADER_ANCHORS.items():
            if text in header_idx:
                cols[key] = header_idx[text] + delta
        # Bereich/Linien-Name liegt relativ zur Medium-Spalte (kein eigener
        # Header in der Bereichs-/Linien-Tabelle vorhanden).
        cols["TOPO_NAME"] = cols["MEDIUM"] + 2

        if any(cols[k] != v for k, v in (
            ("KEY", COL_KEY), ("MEDIUM", COL_MEDIUM), ("PRODUCT", COL_PRODUCT),
        )):
            logger.info(f"XLSX-Spalten abweichend vom Referenzlayout erkannt: {cols}")
        return cols

    def detect_report_type(self, filepath: str) -> str:
        """
        Erkennt den Typ eines ETS6-XLSX-Reports anhand der Kopfzeilen.

        Gibt 'topology' oder 'ga_report' zurück.
        Prüft die ersten 10 Zeilen auf den Report-Titel in Spalte 1.
        """
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl wird für XLSX-Import benoetigt.")
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        report_type = "topology"
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            if not row:
                continue
            title = row[1] if len(row) > 1 and row[1] is not None else ""
            if "gruppenadressen" in str(title).lower():
                report_type = "ga_report"
                break
        wb.close()
        return report_type

    def import_xlsx(self, filepath: str) -> Topology:
        """
        Importiert einen ETS6-Topologie-Report (FA-511).

        Zeilen-Typen werden anhand von Spalte 6 unterschieden:
          - '0'                 → IP-Backbone
          - einzelne Zahl      → Bereich (FA-513)
          - 'B.L'              → Linie  (FA-513)
          - 'B.L.-'            → Spannungsversorgung (FA-516)
          - 'B.L.T'            → Busteilnehmer/Gerät (FA-514)
          - Python-int         → Kommunikationsobjekt (FA-515)
          - Spalte 18 gesetzt  → Geräte-Subzeile (Einbauort)
          - Spalte 38 gesetzt  → KO-Fortsetzungszeile (weitere GAs)
        """
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"XLSX-Datei nicht gefunden: {filepath}")

        if not HAS_OPENPYXL:
            raise ImportError("openpyxl wird für XLSX-Import benoetigt.")

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        topology = Topology()
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            logger.warning(f"Leere XLSX-Datei: {filepath}")
            return topology

        cols = self._resolve_topology_columns(rows)

        # Metadaten aus Kopfzeilen extrahieren (FA-512)
        meta = self._extract_metadata(rows)
        if meta:
            logger.info(
                f"Projektname: {meta.get('project_name', '?')} | "
                f"Druckdatum: {meta.get('print_date', '?')}"
            )
        topology.metadata = meta  # type: ignore[attr-defined]

        current_area: Optional[Area] = None
        current_line: Optional[Line] = None
        current_device: Optional[Device] = None

        for row in rows:
            if not row:
                continue

            key_raw = row[cols["KEY"]] if cols["KEY"] < len(row) else None

            # --- Backbone (Spalte 6 = String '0', nicht Integer 0 = KO-Nr.) ---
            if key_raw == "0":
                medium = _cell(row, cols["MEDIUM"])
                if "IP" in medium.upper():
                    topology.backbone_type = "IP"
                continue

            # --- Bereich erkennen (Spalte 6 = einstellige/-zweistellige Ganzzahl als String) ---
            if isinstance(key_raw, str) and re.match(r"^\d{1,2}$", key_raw.strip()):
                area = self._parse_area_row(row, cols)
                if area:
                    current_area = area
                    topology.areas.append(current_area)
                    current_line = None
                    current_device = None
                continue

            # --- Linie erkennen (Spalte 6 = 'B.L') ---
            if isinstance(key_raw, str) and _LINE_ADDR_RE.match(key_raw.strip()):
                if current_area:
                    line = self._parse_line_row(row, current_area, cols)
                    if line:
                        current_line = line
                        current_area.lines.append(current_line)
                        current_device = None
                continue

            # --- Spannungsversorgung erkennen (Spalte 6 = 'B.L.-') ---
            if isinstance(key_raw, str) and _POWER_SUPPLY_RE.match(key_raw.strip()):
                if current_line:
                    dev = Device(
                        physical_address=key_raw.strip(),
                        manufacturer=_cell(row, cols["MEDIUM"]),
                        order_number=_cell(row, cols["ORDER_NR"]),
                        product=_cell(row, cols["PRODUCT"]),
                        application_program=_cell(row, cols["APP_PROG"]),
                        device_type="power_supply",
                    )
                    current_line.devices.append(dev)
                    current_device = dev
                continue

            # --- Busteilnehmer erkennen (Spalte 6 = 'B.L.T') ---
            if isinstance(key_raw, str) and _PHYS_ADDR_RE.match(key_raw.strip()):
                if current_line:
                    device = self._parse_device_row(row, cols)
                    current_line.devices.append(device)
                    current_device = device
                continue

            # --- Kommunikationsobjekt (Spalte 6 = Python-int) ---
            if isinstance(key_raw, int):
                ko = self._parse_ko_row(row, key_raw, cols)
                if ko and current_device:
                    current_device.communication_objects.append(ko)
                continue

            # --- Geräte-Subzeile: Einbauort (Spalte 18 gesetzt, Spalte 6 leer) ---
            loc = _cell(row, cols["LOCATION"])
            if loc and key_raw is None and current_device:
                current_device.installation_location = loc
                continue

            # --- KO-Fortsetzungszeile: weitere verbundene GAs (Spalte 38 gesetzt) ---
            ga_cont = _cell(row, cols["GA"])
            if (ga_cont and key_raw is None
                    and current_device
                    and current_device.communication_objects):
                last_ko = current_device.communication_objects[-1]
                for ga in self._extract_gas(ga_cont):
                    if ga not in last_ko.connected_gas:
                        last_ko.connected_gas.append(ga)

        # Geräteanzahl berechnen und Grenzwerte prüfen (FA-518)
        for area in topology.areas:
            for line in area.lines:
                line.update_device_count()
                if line.device_count > 100:
                    logger.error(
                        f"Linie {line.coupler_address}: "
                        f"{line.device_count} Geräte (max. empfohlen: 100)"
                    )
                elif line.device_count > 85:
                    logger.warning(
                        f"Linie {line.coupler_address}: "
                        f"{line.device_count} Geräte (Empfehlung: max. 85)"
                    )

        total_lines = sum(len(a.lines) for a in topology.areas)
        total_devices = sum(
            len(l.devices)
            for a in topology.areas
            for l in a.lines
        )
        logger.info(
            f"XLSX importiert: {len(topology.areas)} Bereiche, "
            f"{total_lines} Linien, {total_devices} Geräte  [{filepath}]"
        )
        return topology

    # ------------------------------------------------------------------
    # Hilfsmethoden
    # ------------------------------------------------------------------

    def _extract_metadata(self, rows: list) -> dict:
        """
        Extrahiert Projektmetadaten aus den Kopfzeilen (FA-512).

        Sucht das Label (z.B. "Projekt:") an beliebiger Spaltenposition in
        der Zeile und nimmt die erste nicht-leere Zelle danach als Wert --
        feste Spaltenindizes sind nicht zuverlässig, da ETS6 je nach
        Exportkonfiguration unterschiedlich viele Spalten einfügt.
        """
        meta: dict = {}
        label_map = {
            "projekt":     "project_name",
            "project":     "project_name",
            "startdatum":  "start_date",
            "importdatum": "import_date",
            "druckdatum":  "print_date",
            "druckzeit":   "print_time",
        }
        for row in rows[:20]:
            if not row:
                continue
            for idx, cell in enumerate(row):
                if not isinstance(cell, str):
                    continue
                label = cell.strip().lower().rstrip(":")
                if label not in label_map or label_map[label] in meta:
                    continue
                for val in row[idx + 1:]:
                    if val is not None and str(val).strip():
                        meta[label_map[label]] = str(val).strip()
                        break
        return meta

    def _parse_area_row(self, row: tuple, cols: dict[str, int]) -> Optional[Area]:
        """Erkennt und liest eine Bereich-Zeile (FA-513)."""
        key = _cell(row, cols["KEY"])
        try:
            area_num = int(key)
        except ValueError:
            return None
        if area_num == 0:
            return None  # Backbone, kein Bereich
        medium = _cell(row, cols["MEDIUM"])
        name = _cell(row, cols["TOPO_NAME"]) or f"Bereich {area_num}"
        return Area(
            area_number=area_num,
            name=name,
            coupler_address=f"{area_num}.0.0",
            backbone_type="IP" if "IP" in medium.upper() else "TP",
        )

    def _parse_line_row(self, row: tuple, area: Area, cols: dict[str, int]) -> Optional[Line]:
        """Erkennt und liest eine Linien-Zeile (FA-513)."""
        key = _cell(row, cols["KEY"])
        m = _LINE_ADDR_RE.match(key)
        if not m:
            return None
        line_num = int(m.group(2))
        name = _cell(row, cols["TOPO_NAME"]) or f"Linie {line_num}"
        return Line(
            line_number=line_num,
            name=name,
            coupler_address=f"{area.area_number}.{line_num}.0",
        )

    def _parse_device_row(self, row: tuple, cols: dict[str, int]) -> Device:
        """Liest eine Busteilnehmer-Zeile (FA-514)."""
        addr = _cell(row, cols["KEY"])
        m = _PHYS_ADDR_RE.match(addr)
        participant = int(m.group(3)) if m else -1

        product = _cell(row, cols["PRODUCT"])

        # Koppler: Teilnehmer-Adresse == 0 (FA-517)
        if participant == 0:
            dev_type = "coupler"
        elif participant < 0:
            dev_type = "unknown"
        else:
            dev_type = self._classify_device(addr, product)

        return Device(
            physical_address=addr,
            manufacturer=_cell(row, cols["MEDIUM"]),
            order_number=_cell(row, cols["ORDER_NR"]),
            product=product,
            application_program=_cell(row, cols["APP_PROG"]),
            device_type=dev_type,
        )

    def _parse_ko_row(
        self, row: tuple, obj_num: int, cols: dict[str, int],
    ) -> Optional[CommunicationObject]:
        """Liest eine Kommunikationsobjekt-Zeile (FA-515)."""
        name = _cell(row, cols["KO_NAME"])
        # Kopfzeilen der KO-Tabelle (z.B. "Name", " #") uebergehen
        if name.lower() in ("name", "beschreibung", ""):
            return None

        ga_cell = _cell(row, cols["GA"])
        connected_gas = self._extract_gas(ga_cell) if ga_cell else []

        return CommunicationObject(
            object_number=obj_num,
            name=name,
            object_function=_cell(row, cols["KO_FUNC"]),
            priority=_cell(row, cols["KO_PRIO"]) or "Niedrig",
            flags=_cell(row, cols["KO_FLAGS"]),
            data_type=_cell(row, cols["KO_DT"]),
            connected_gas=connected_gas,
        )

    def _extract_gas(self, ga_text: str) -> list[str]:
        """
        Extrahiert alle GA-Adressen (Format H/M/S) aus einem Zellentext (FA-519).

        Beispiel: '3/0/65 L.OG.01.01_ea   ( Licht )  0/0/100 3/4/9'
        → ['3/0/65', '0/0/100', '3/4/9']
        """
        return _GA_RE.findall(ga_text)

    def _classify_device(self, address: str, product: str = "") -> str:
        """Klassifiziert ein Gerät anhand Produktname (bevorzugt) und Adresse.

        Reine Adress-Heuristik (Teilnehmer <=100 -> Aktor, <=199 -> Sensor)
        liegt oft falsch, z.B. bei Tastern mit niedriger Adresse. Der
        Produktname liefert ein zuverlässigeres Signal (analog
        knxproj_import_service._infer_device_type) und wird deshalb zuerst
        geprüft; die Adress-Heuristik bleibt Fallback, wenn keine
        Schlüsselwörter zutreffen oder kein Produktname vorliegt.
        """
        m = _PHYS_ADDR_RE.match(address)
        if not m:
            return "unknown"
        participant = int(m.group(3))
        if participant == 0:
            return "coupler"

        name_lower = (product or "").lower()
        if any(kw in name_lower for kw in _INFRA_KEYWORDS):
            return "other"
        if any(kw in name_lower for kw in _SENSOR_KEYWORDS):
            return "sensor"
        if any(kw in name_lower for kw in _ACTOR_KEYWORDS):
            return "actor"

        if participant <= 100:
            return "actor"
        if participant <= 199:
            return "sensor"
        return "unknown"

    def derive_building_structure(
        self, filepath: str, ga_structure: "GroupAddressStructure | None" = None,
    ) -> Areal:
        """
        Leitet Gebäudestruktur aus den GA-Namen im XLSX-Topologie-Report ab.

        Scannt alle GA-Zellen nach dem Muster '{Gewerk}.{Stockwerk}.{RaumNr}...'
        und erstellt daraus Stockwerke und Räume (FA-506 analog für XLSX).

        Zusätzlich wird der Einbauort (Spalte S / COL_LOCATION) jedes Geräts
        ausgelesen. Da der Einbauort das Format 'NN  Raumname' hat und in den
        KO-Zeilen desselben Geräts der Stockwerk-Code vorkommt, kann die
        Zuordnung (Stockwerk, RaumNr) → echter Raumname aus Spalte S hergestellt
        werden. Dieser Name hat Vorrang vor dem aus GA-Beschreibungen abgeleiteten.

        Manche Projekte benennen GAs ohne Stockwerk-Buchstaben, nach der
        EFH-Konvention '{Gewerk}.{Stockwerk}{RaumNr}.{Element}' (z.B.
        'L.004.1_ea' = Stockwerk 0, Raum 04). Liefert das Buchstaben-Muster
        keine Treffer, wird auf dieses Ziffern-Muster zurückgegriffen; der
        Stockwerksname wird dabei aus der passenden Hauptgruppe von
        `ga_structure` übernommen (Hauptgruppe = Stockwerk-Ziffer + 1), falls
        übergeben, sonst generisch als "Stockwerk N" benannt.

        Gibt ein Areal mit Gebäude > Flügel > Stockwerke > Wohnungen > Räume zurück.
        """
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"XLSX-Datei nicht gefunden: {filepath}")
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl wird für XLSX-Import benoetigt.")

        # floor_rooms: {floor_code -> {room_nr -> [ga-beschreibungen]}}
        floor_rooms: dict[str, dict[str, list[str]]] = {}
        # col_s_names: {(floor_code, room_nr) -> echter Raumname aus Spalte S}
        col_s_names: dict[tuple[str, str], str] = {}
        # digit_floor_rooms: {stockwerk_ziffer -> {room_nr -> [ga-beschreibungen]}}
        digit_floor_rooms: dict[str, dict[str, list[str]]] = {}

        _LOC_RE = re.compile(r"^(\d{2})\s+(.+)$")

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        cols = self._resolve_topology_columns(rows)
        current_location: str = ""  # Raumname aus Spalte S des aktuellen Geräts

        for row in rows:
            if not row:
                continue
            key_raw = row[cols["KEY"]] if cols["KEY"] < len(row) else None

            # Neue Gerätezeile → Einbauort zurücksetzen
            if isinstance(key_raw, str) and (
                _PHYS_ADDR_RE.match(key_raw.strip())
                or _POWER_SUPPLY_RE.match(key_raw.strip())
            ):
                current_location = ""
                continue

            # Geräte-Subzeile: Einbauort aus Spalte S lesen
            if key_raw is None:
                loc_raw = row[cols["LOCATION"]] if cols["LOCATION"] < len(row) else None
                if loc_raw is not None:
                    loc_m = _LOC_RE.match(str(loc_raw).strip())
                    if loc_m:
                        current_location = loc_m.group(2).strip()

            # KO-Zeile: GA-Namen auswerten und Stockwerk/Raum-Mapping aufbauen
            if isinstance(key_raw, int):
                if cols["GA"] < len(row) and row[cols["GA"]] is not None:
                    text = str(row[cols["GA"]])
                    for m in _GA_FLOOR_ROOM_RE.finditer(text):
                        floor_code = m.group(1).upper()
                        room_nr = m.group(2)
                        desc = m.group(3).strip()

                        if floor_code not in floor_rooms:
                            floor_rooms[floor_code] = {}
                        if room_nr not in floor_rooms[floor_code]:
                            floor_rooms[floor_code][room_nr] = []
                        if desc and desc not in floor_rooms[floor_code][room_nr]:
                            floor_rooms[floor_code][room_nr].append(desc)

                        # Spalte-S-Namen mit (Stockwerk, RaumNr) verknüpfen
                        fk = (floor_code, room_nr)
                        if current_location and fk not in col_s_names:
                            col_s_names[fk] = current_location

                    for m in _GA_DIGIT_ROOM_RE.finditer(text):
                        digit = m.group(1)
                        room_nr = m.group(2)
                        desc = (m.group(3) or "").strip()

                        if digit not in digit_floor_rooms:
                            digit_floor_rooms[digit] = {}
                        if room_nr not in digit_floor_rooms[digit]:
                            digit_floor_rooms[digit][room_nr] = []
                        if desc and desc not in digit_floor_rooms[digit][room_nr]:
                            digit_floor_rooms[digit][room_nr].append(desc)

        if not floor_rooms and digit_floor_rooms:
            return self._build_areal_from_digit_floors(digit_floor_rooms, ga_structure)

        if not floor_rooms:
            logger.warning("Keine Gebäudestruktur aus XLSX ableitbar (keine GA-Namen gefunden).")
            areal = Areal(name="Importiertes Projekt")
            building = Building(name="Gebäude")
            wing = Wing(name="Hauptgebäude")
            building.wings.append(wing)
            areal.buildings.append(building)
            return areal

        logger.info(
            f"Raumname aus Spalte S zugeordnet für {len(col_s_names)} Stockwerk/Raum-Kombinationen"
        )

        # Gebäude-Hierarchie aufbauen
        areal = Areal(name="")
        building = Building(name="Gebäude")
        wing = Wing(name="Hauptgebäude")

        # Stockwerke sortieren: UG < EG < OG < 1.OG < 2.OG < 3.OG < DG
        sort_order = {code: i for i, code in enumerate(
            ["UG", "EG", "OG", "1OG", "1.OG", "2OG", "2.OG", "3OG", "3.OG", "DG"]
        )}
        sorted_floors = sorted(
            floor_rooms.keys(),
            key=lambda c: sort_order.get(c, 99),
        )

        for floor_code in sorted_floors:
            rooms_dict = floor_rooms[floor_code]
            full_name = STANDARD_FLOOR_NAMES.get(floor_code, floor_code)
            hg_number = FLOOR_TO_MAIN_GROUP.get(floor_code, -1)
            floor = Floor(
                name=full_name,
                short_code=floor_code,
                main_group_number=hg_number,
            )
            apartment = Apartment(name=floor_code)
            floor.apartments.append(apartment)

            for room_nr in sorted(rooms_dict.keys()):
                fk = (floor_code, room_nr)
                if fk in col_s_names:
                    room_name = col_s_names[fk]
                else:
                    room_name = self._best_room_name(rooms_dict[room_nr], room_nr)
                room = Room(number=room_nr, name=room_name)
                apartment.rooms.append(room)

            wing.floors.append(floor)
            logger.info(
                f"Stockwerk {floor_code} ({full_name}): {len(rooms_dict)} Räume"
            )

        building.wings.append(wing)
        areal.buildings.append(building)

        total_rooms = sum(len(r) for r in floor_rooms.values())
        logger.info(
            f"Gebäudestruktur abgeleitet: {len(floor_rooms)} Stockwerke, "
            f"{total_rooms} Räume  [{filepath}]"
        )
        return areal

    def _build_areal_from_digit_floors(
        self,
        digit_floor_rooms: dict[str, dict[str, list[str]]],
        ga_structure: "GroupAddressStructure | None",
    ) -> Areal:
        """
        Baut ein Areal aus der Ziffern-Stockwerk-Konvention ('{Gewerk}.{D}{nn}.{Elem}').

        Stockwerksname wird, falls verfügbar, aus der ETS-Hauptgruppe
        übernommen (Hauptgruppe = Stockwerk-Ziffer + 1) -- diese trägt im
        GA-Report meist den echten Stockwerksnamen (z.B. "Erdgeschoss").
        Ohne GA-Struktur wird generisch "Stockwerk N" verwendet.
        """
        hg_names: dict[int, str] = {}
        if ga_structure is not None:
            for hg in ga_structure.main_groups:
                if hg.name and not hg.name.lower().startswith("hauptgruppe"):
                    hg_names[hg.number] = hg.name

        areal = Areal(name="")
        building = Building(name="Gebäude")
        wing = Wing(name="Hauptgebäude")

        for digit in sorted(digit_floor_rooms.keys(), key=int):
            rooms_dict = digit_floor_rooms[digit]
            hg_number = int(digit) + 1
            full_name = hg_names.get(hg_number, f"Stockwerk {digit}")
            short_code = f"D{digit}"
            floor = Floor(
                name=full_name,
                short_code=short_code,
                main_group_number=hg_number,
            )
            apartment = Apartment(name=short_code)
            floor.apartments.append(apartment)

            for room_nr in sorted(rooms_dict.keys()):
                room_name = self._best_room_name(rooms_dict[room_nr], room_nr)
                room = Room(number=room_nr, name=room_name)
                apartment.rooms.append(room)

            wing.floors.append(floor)
            logger.info(f"Stockwerk {short_code} ({full_name}): {len(rooms_dict)} Räume")

        building.wings.append(wing)
        areal.buildings.append(building)

        total_rooms = sum(len(r) for r in digit_floor_rooms.values())
        logger.info(
            f"Gebäudestruktur (Ziffern-Konvention) abgeleitet: "
            f"{len(digit_floor_rooms)} Stockwerke, {total_rooms} Räume"
        )
        return areal

    def _best_room_name(self, descriptions: list[str], fallback_nr: str) -> str:
        """Waehlt den besten Raumnamen aus einer Liste von GA-Beschreibungen."""
        if not descriptions:
            return f"Raum {fallback_nr}"
        # Bevorzuge Beschreibungen, die nicht rein technisch/direktional sind
        candidates = [
            d for d in descriptions
            if d.lower().split()[0] not in _NON_ROOM_WORDS and len(d) >= 4
        ]
        if not candidates:
            candidates = descriptions
        # Kuerzeste sinnvolle Beschreibung nehmen (oft der eigentliche Raumname)
        best = min(candidates, key=len)
        # Auf 30 Zeichen begrenzen
        return best[:30] if len(best) <= 30 else best[:27] + "..."

    def extract_group_addresses(self, filepath: str) -> GroupAddressStructure:
        """
        Extrahiert Gruppenadressen aus dem ETS6-Topologie-Report (XLSX) (FA-519).

        Scannt alle GA-Zellen (Spalte 38) und baut daraus eine
        GroupAddressStructure auf, geordnet nach Haupt- und Mittelgruppe.
        Bereits bekannte Adressen werden nicht doppelt eingefügt.
        """
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"XLSX-Datei nicht gefunden: {filepath}")
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl wird für XLSX-Import benoetigt.")

        # Eindeutige GA-Eintraege sammeln: addr_str -> designation-Text
        ga_map: dict[str, str] = {}

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        cols = self._resolve_topology_columns(rows)
        ga_col = cols["GA"]
        for row in rows:
            if not row or len(row) <= ga_col or row[ga_col] is None:
                continue
            text = str(row[ga_col]).strip()
            if not text:
                continue
            for addr, designation in self._parse_ga_entries(text):
                if addr not in ga_map:
                    ga_map[addr] = designation

        if not ga_map:
            logger.warning("Keine Gruppenadressen aus XLSX extrahierbar.")
            return GroupAddressStructure()

        structure = GroupAddressStructure()
        hg_index: dict[int, MainGroup] = {}
        mg_index: dict[tuple[int, int], MiddleGroup] = {}

        for addr_str, designation in sorted(ga_map.items(), key=lambda x: _ga_sort_key(x[0])):
            parts = addr_str.split("/")
            if len(parts) != 3:
                continue
            try:
                hg_num = int(parts[0])
                mg_num = int(parts[1])
                sub_num = int(parts[2])
            except ValueError:
                continue

            if hg_num not in hg_index:
                hg = MainGroup(number=hg_num, name=f"Hauptgruppe {hg_num}")
                hg_index[hg_num] = hg
                structure.main_groups.append(hg)

            mg_key = (hg_num, mg_num)
            if mg_key not in mg_index:
                mg = MiddleGroup(number=mg_num, name=f"Mittelgruppe {mg_num}")
                mg_index[mg_key] = mg
                hg_index[hg_num].middle_groups.append(mg)

            ga = GroupAddress(
                main_group=hg_num,
                middle_group=mg_num,
                sub_group=sub_num,
                designation=designation,
            )
            self._parse_xlsx_designation(ga, designation)
            mg_index[mg_key].group_addresses.append(ga)

        structure.source = "topology"
        logger.info(
            f"XLSX GA-Extraktion: {len(ga_map)} Gruppenadressen, "
            f"{len(hg_index)} Hauptgruppen, {len(mg_index)} Mittelgruppen"
            f"  [{filepath}]"
        )
        return structure

    def _parse_ga_entries(self, text: str) -> list[tuple[str, str]]:
        """
        Extrahiert (Adresse, Bezeichnung)-Paare aus einem GA-Zellentext.

        Beispiel: '3/0/65 LD.EG.05.01_ea ( Licht )  0/0/100'
        -> [('3/0/65', 'LD.EG.05.01_ea ( Licht )'), ('0/0/100', '')]
        """
        matches = list(_GA_RE.finditer(text))
        entries: list[tuple[str, str]] = []
        for i, m in enumerate(matches):
            addr = m.group()
            start = m.end()
            end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
            designation = text[start:end].strip()
            entries.append((addr, designation))
        return entries

    def _parse_xlsx_designation(self, ga: GroupAddress, designation: str):
        """
        Extrahiert Gewerk-Code und Beschreibung aus einer XLSX-GA-Bezeichnung.

        Erwartet das Muster 'GEWERK.ETAGE.RAUM.ELEM_suffix ( Beschreibung )',
        z.B. 'LD.EG.05.01_ea ( Licht )'.
        """
        if not designation:
            return
        # Beschreibung aus Klammern: 'LD.EG.05.01_ea ( Licht )' -> 'Licht'
        paren_m = re.search(r'\(\s*([^)]+)\)', designation)
        if paren_m:
            ga.description = paren_m.group(1).strip()
            base = designation[:paren_m.start()].strip()
        else:
            base = designation.strip()
        # Gewerk-Code: erstes Segment vor '.' oder '_'
        if '.' in base:
            ga.gewerk_code = base.split('.')[0].upper()
        elif '_' in base:
            ga.gewerk_code = base.split('_')[0].upper()

    def import_ga_report(self, filepath: str) -> GroupAddressStructure:
        """
        Importiert einen ETS6 Gruppenadress-Report (XLSX) (FA-519b).

        Zeilentypen werden anhand von Spalte 4 unterschieden:
          - Einstellige/-zweistellige Ganzzahl  → Hauptgruppe (col 8 = Name)
          - 'H/M'                               → Mittelgruppe (col 8 = Name)
          - 'H/M/S'                             → Gruppenadresse (col 8 = Name, col 22 = Datentyp)
          - Physikalische Adresse / KO / Header → wird übersprungen
        """
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"XLSX-Datei nicht gefunden: {filepath}")
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl wird für XLSX-Import benoetigt.")

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        structure = GroupAddressStructure()
        hg_index: dict[int, MainGroup] = {}
        mg_index: dict[tuple[int, int], MiddleGroup] = {}

        current_hg: Optional[MainGroup] = None
        current_mg: Optional[MiddleGroup] = None

        meta: dict = {}
        row_count = 0

        for row in ws.iter_rows(values_only=True):
            if not row:
                continue

            # Metadaten aus Kopfzeilen (erste 20 Zeilen)
            if row_count < 20:
                label_raw = row[COL_GAR_META_LABEL] if COL_GAR_META_LABEL < len(row) else None
                if label_raw is not None:
                    label = str(label_raw).strip().lower().rstrip(":")
                    label_map = {
                        "projekt": "project_name", "project": "project_name",
                        "startdatum": "start_date", "importdatum": "import_date",
                        "druckdatum": "print_date", "druckzeit": "print_time",
                    }
                    if label in label_map:
                        val_raw = row[COL_GAR_META_VALUE] if COL_GAR_META_VALUE < len(row) else None
                        if val_raw is not None:
                            meta[label_map[label]] = str(val_raw).strip()
            row_count += 1

            addr_raw = row[COL_GAR_ADDR] if COL_GAR_ADDR < len(row) else None
            if addr_raw is None:
                continue
            addr = str(addr_raw).strip()

            # Hauptgruppe
            if _GAR_HG_RE.match(addr):
                try:
                    hg_num = int(addr)
                except ValueError:
                    continue
                name_raw = row[COL_GAR_NAME] if COL_GAR_NAME < len(row) else None
                name = str(name_raw).strip() if name_raw else f"Hauptgruppe {hg_num}"
                if hg_num not in hg_index:
                    hg = MainGroup(number=hg_num, name=name)
                    hg_index[hg_num] = hg
                    structure.main_groups.append(hg)
                current_hg = hg_index[hg_num]
                current_mg = None
                continue

            # Mittelgruppe
            if _GAR_MG_RE.match(addr):
                parts = addr.split("/")
                try:
                    hg_num, mg_num = int(parts[0]), int(parts[1])
                except ValueError:
                    continue
                name_raw = row[COL_GAR_NAME] if COL_GAR_NAME < len(row) else None
                name = str(name_raw).strip() if name_raw else f"Mittelgruppe {mg_num}"
                mg_key = (hg_num, mg_num)
                if mg_key not in mg_index:
                    # Hauptgruppe sicherstellen (defensiv)
                    if hg_num not in hg_index:
                        hg = MainGroup(number=hg_num, name=f"Hauptgruppe {hg_num}")
                        hg_index[hg_num] = hg
                        structure.main_groups.append(hg)
                    mg = MiddleGroup(number=mg_num, name=name)
                    mg_index[mg_key] = mg
                    hg_index[hg_num].middle_groups.append(mg)
                current_mg = mg_index[mg_key]
                continue

            # Gruppenadresse
            if _GAR_GA_RE.match(addr):
                parts = addr.split("/")
                try:
                    hg_num, mg_num, sub_num = int(parts[0]), int(parts[1]), int(parts[2])
                except ValueError:
                    continue
                name_raw = row[COL_GAR_NAME] if COL_GAR_NAME < len(row) else None
                designation = str(name_raw).strip() if name_raw else ""
                dtype_raw = row[COL_GAR_DTYPE] if COL_GAR_DTYPE < len(row) else None
                data_type = str(dtype_raw).strip() if dtype_raw else ""

                # Mittelgruppe sicherstellen (defensiv)
                mg_key = (hg_num, mg_num)
                if mg_key not in mg_index:
                    if hg_num not in hg_index:
                        hg = MainGroup(number=hg_num, name=f"Hauptgruppe {hg_num}")
                        hg_index[hg_num] = hg
                        structure.main_groups.append(hg)
                    mg = MiddleGroup(number=mg_num, name=f"Mittelgruppe {mg_num}")
                    mg_index[mg_key] = mg
                    hg_index[hg_num].middle_groups.append(mg)

                ga = GroupAddress(
                    main_group=hg_num,
                    middle_group=mg_num,
                    sub_group=sub_num,
                    designation=designation,
                    datapoint_type=data_type,
                )
                self._parse_xlsx_designation(ga, designation)
                mg_index[mg_key].group_addresses.append(ga)
                continue

        wb.close()

        if meta:
            logger.info(
                f"GA-Report Projektname: {meta.get('project_name', '?')} | "
                f"Druckdatum: {meta.get('print_date', '?')}"
            )

        total_gas = sum(
            len(mg.group_addresses)
            for hg in structure.main_groups
            for mg in hg.middle_groups
        )
        structure.source = "ga_report"
        logger.info(
            f"GA-Report importiert: {len(hg_index)} Hauptgruppen, "
            f"{len(mg_index)} Mittelgruppen, {total_gas} Gruppenadressen  [{filepath}]"
        )
        return structure

    def merge_with_csv(self, topology: Topology, ga_structure) -> int:
        """
        Fuehrt Topologie-Report und CSV-GA-Struktur zusammen (FA-520).

        Verknuepft Kommunikationsobjekte mit Gruppenadressen anhand
        der GA-Adresse und gibt die Anzahl der Verknuepfungen zurück.
        """
        ga_index: dict[str, object] = {}
        for ga in ga_structure.all_addresses():
            ga_index[ga.address] = ga

        linked = 0
        for area in topology.areas:
            for line in area.lines:
                for device in line.devices:
                    for ko in device.communication_objects:
                        for ga_addr in ko.connected_gas:
                            if ga_addr in ga_index:
                                linked += 1

        logger.info(f"Zusammenfuehrung: {linked} GA-Verknuepfungen hergestellt")
        return linked

    def extract_device_locations(self, filepath: str) -> dict[str, dict]:
        """
        Liest Einbauort-Informationen aus einem ETS6 GA-Report (XLSX).

        Gibt ein Dict zurück: {physikalische_adresse → info}

        'info' hat zwei Formen:
          {"type": "room",      "room_nr": "04", "room_name": "Schlafen"}
          {"type": "verteiler", "vt_type": "UV2", "vt_name": "Steigzone"}
        """
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"XLSX-Datei nicht gefunden: {filepath}")
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl wird für XLSX-Import benoetigt.")

        _VT_RE = re.compile(r"^([A-Z0-9]+)\s*\(\s*(.+?)\s*\)$")

        result: dict[str, dict] = {}
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        for row in ws.iter_rows(values_only=True):
            if not row or COL_GAR_ADDR >= len(row) or row[COL_GAR_ADDR] is None:
                continue
            addr = str(row[COL_GAR_ADDR]).strip()
            if not _PHYS_ADDR_RE.match(addr):
                continue
            loc_raw = row[COL_GAR_LOCATION] if COL_GAR_LOCATION < len(row) else None
            if loc_raw is None:
                continue
            loc = str(loc_raw).strip()

            m_room = _EINBAUORT_ROOM_RE.match(loc)
            if m_room:
                result[addr] = {
                    "type": "room",
                    "room_nr": m_room.group(1),
                    "room_name": m_room.group(2).strip(),
                }
                continue

            m_vt = _VT_RE.match(loc)
            if m_vt:
                result[addr] = {
                    "type": "verteiler",
                    "vt_type": m_vt.group(1).strip(),
                    "vt_name": m_vt.group(2).strip(),
                }

        wb.close()
        logger.info(
            f"extract_device_locations: {sum(1 for v in result.values() if v['type'] == 'room')} "
            f"Raum-Zuordnungen, "
            f"{sum(1 for v in result.values() if v['type'] == 'verteiler')} "
            f"Verteiler-Zuordnungen  [{filepath}]"
        )
        return result

    def extract_device_notes(self, filepath: str) -> dict[str, str]:
        """
        Liest das Feld 'Installations-Hinweise' pro Gerät aus dem ETS6-
        Topologie-Report (XLSX).

        ETS6 stapelt unterhalb jeder Geräte-Hauptzeile bis zu vier Subzeilen
        (Beschreibung / Seriennummer / Kommentar / Installations-Hinweise),
        alle in Spalte 6 -- derselben Spalte wie die physikalische Adresse.
        Leere Felder erzeugen keine eigene Zeile, daher liegt der Hinweistext
        je Gerät auf unterschiedlicher Zeilenposition; gesucht wird die erste
        Subzeile mit Text in Spalte 6, die nicht wie eine Seriennummer aussieht.

        Installateure tragen hier oft die GA-Kurzbezeichnung ein, die das
        Gerät bedient (z.B. 'L.100.1' oder 'S.203.1_LD.203.1'). Das ist bei
        Aktoren im Verteiler eine zuverlässigere Quelle für den funktional
        bedienten Raum als der physische Einbauort.

        Gibt {physikalische_adresse -> Hinweistext} zurück (nur nicht-leere,
        nicht Seriennummer-artige Werte).
        """
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"XLSX-Datei nicht gefunden: {filepath}")
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl wird für XLSX-Import benoetigt.")

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        cols = self._resolve_topology_columns(rows)
        key_col = cols["KEY"]
        notes: dict[str, str] = {}
        current_addr: Optional[str] = None

        for row in rows:
            key_raw = row[key_col] if key_col < len(row) else None

            if key_raw == "0":
                current_addr = None  # Backbone
                continue
            if isinstance(key_raw, str) and re.match(r"^\d{1,2}$", key_raw.strip()):
                current_addr = None  # Bereich
                continue
            if isinstance(key_raw, str) and _LINE_ADDR_RE.match(key_raw.strip()):
                current_addr = None  # Linie
                continue
            if isinstance(key_raw, str) and _POWER_SUPPLY_RE.match(key_raw.strip()):
                current_addr = None  # Spannungsversorgung
                continue
            if isinstance(key_raw, str) and _PHYS_ADDR_RE.match(key_raw.strip()):
                current_addr = key_raw.strip()
                continue
            if isinstance(key_raw, int):
                continue  # KO-Zeile

            if current_addr and isinstance(key_raw, str):
                text = key_raw.strip()
                if (
                    text
                    and not _SERIAL_NUMBER_RE.match(text)
                    and current_addr not in notes
                ):
                    notes[current_addr] = text

        logger.info(
            f"extract_device_notes: {len(notes)} Installations-Hinweise gefunden  [{filepath}]"
        )
        return notes

    @staticmethod
    def _room_id_for_designation(
        designation: str, room_id_by_floor_nr: dict[tuple[str, str], str],
    ) -> Optional[str]:
        """Löst eine GA-Bezeichnung (oder einen Installations-Hinweis-Token)
        über die Buchstaben- oder Ziffern-Stockwerkskonvention in eine
        Room-ID auf (gemeinsame Logik für mehrere Raum-Zuordnungsquellen)."""
        m = _GA_FLOOR_ROOM_SIMPLE_RE.match(designation)
        if m:
            rid = room_id_by_floor_nr.get((m.group(1).upper(), m.group(2)))
            if rid:
                return rid
        m_digit = _GA_DIGIT_ROOM_SIMPLE_RE.match(designation)
        if m_digit:
            return room_id_by_floor_nr.get((f"D{m_digit.group(1)}", m_digit.group(2)))
        return None

    def link_rooms_to_lines(
        self,
        topology: Topology,
        ga_structure,
        areal,
        device_locations: dict[str, dict] | None = None,
        device_notes: dict[str, str] | None = None,
    ) -> int:
        """
        Verknüpft Topologie-Linien mit Räumen aus der Gebäudestruktur.

        Drei Quellen werden kombiniert, in dieser Priorität:

        1. Installations-Hinweise (device_notes aus extract_device_notes):
           Freitext-Notiz des Installateurs, z.B. 'L.100.1' oder
           'S.203.1_LD.203.1'. Explizitester Hinweis auf den funktional
           bedienten Raum, sofern gepflegt -- höchste Priorität.

        2. GA-Bezeichnungen der Kommunikationsobjekte (funktionaler Raum):
           'LDA.OG.00.01_ea' bzw. 'L.004.1_ea' → Stockwerk + Raum-Nr → Room-ID.
           Spiegelt den Raum wider, den die GA tatsächlich steuert. Wichtig
           bei Aktoren, deren Einbauort (s.u.) i.d.R. der Verteiler ist,
           nicht der bediente Raum.

        3. Einbauort (device_locations aus extract_device_locations):
           "04 Schlafen" → room_nr + room_name → Room-ID.
           Fallback für Geräte ohne auswertbare GA-Bezeichnung/Hinweis.
           Liefert bei Aktoren oft den physischen Einbauort (z.B. den
           Verteiler) statt des bedienten Raums und ist daher zuletzt eingestuft.

        Pro Linie werden alle gefundenen Room-IDs in assigned_room_ids eingetragen.
        Pro Gerät wird die häufigste Room-ID als device.room_id gesetzt.

        Gibt die Anzahl der verknüpften (Linie, Raum)-Paare zurück.
        """
        # GA-Adress-Bezeichnungs-Index aufbauen
        ga_designation: dict[str, str] = {}
        for ga in ga_structure.all_addresses():
            if ga.designation:
                ga_designation[ga.address] = ga.designation

        # Raum-Index 1: (floor_code, room_nr) → room_id  (für GA-Bezeichnungs-Fallback)
        room_id_by_floor_nr: dict[tuple[str, str], str] = {}
        # Raum-Index 2: (room_nr, room_name_lower) → room_id  (für Einbauort)
        room_id_by_nr_name: dict[tuple[str, str], str] = {}

        for building in areal.buildings:
            for wing in building.wings:
                for floor in wing.floors:
                    for apartment in floor.apartments:
                        for room in apartment.rooms:
                            room_id_by_floor_nr[(floor.short_code, room.number)] = room.id
                            room_id_by_nr_name[
                                (room.number, room.name.lower().strip())
                            ] = room.id

        if not room_id_by_floor_nr:
            logger.warning("link_rooms_to_lines: Keine Räume in Gebäudestruktur.")
            return 0

        linked_pairs = 0
        via_location = 0
        via_ga = 0
        via_note = 0

        for area in topology.areas:
            for line in area.lines:
                line_room_ids: list[str] = []

                for device in line.devices:
                    room_id = None

                    # ── Quelle 1: Installations-Hinweis (expliziter Installateur-Hinweis) ──
                    note = (device_notes or {}).get(device.physical_address, "")
                    if note:
                        note_room_counts: dict[str, int] = {}
                        for token in re.split(r"[_\s]+", note.strip()):
                            if not token:
                                continue
                            rid = self._room_id_for_designation(token, room_id_by_floor_nr)
                            if rid:
                                note_room_counts[rid] = note_room_counts.get(rid, 0) + 1
                        if note_room_counts:
                            room_id = max(
                                note_room_counts, key=note_room_counts.__getitem__
                            )
                            via_note += 1

                    # ── Quelle 2: GA-Bezeichnungen der KOs (funktionaler Raum) ──
                    if not room_id:
                        device_room_counts: dict[str, int] = {}
                        for ko in device.communication_objects:
                            for ga_addr in ko.connected_gas:
                                designation = ga_designation.get(ga_addr, "")
                                if not designation:
                                    continue
                                rid = self._room_id_for_designation(
                                    designation, room_id_by_floor_nr
                                )
                                if rid:
                                    device_room_counts[rid] = (
                                        device_room_counts.get(rid, 0) + 1
                                    )
                        if device_room_counts:
                            room_id = max(
                                device_room_counts, key=device_room_counts.__getitem__
                            )
                            via_ga += 1

                    # ── Quelle 3: Einbauort aus GA-Report (Fallback, physischer Ort) ──
                    if not room_id and device_locations:
                        loc = device_locations.get(device.physical_address)
                        if loc and loc["type"] == "room":
                            nr = loc["room_nr"]
                            raw_name = loc["room_name"].lower().strip()
                            # Normalisierung: '/' entfernen, mehrfache Leerzeichen kürzen
                            norm_name = re.sub(r"\s*/\s*", " ", raw_name)
                            norm_name = re.sub(r"\s+", " ", norm_name).strip()
                            room_id = room_id_by_nr_name.get(
                                (nr, raw_name)
                            ) or room_id_by_nr_name.get((nr, norm_name))
                            # Teilwort-Fallback: alle Wörter des Einbauorts
                            # müssen im Areal-Raumnamen vorkommen (gleiche Nr.)
                            if not room_id:
                                words = [w for w in norm_name.split() if len(w) > 2]
                                if words:
                                    for (rnr, rname_lc), rid in room_id_by_nr_name.items():
                                        if rnr == nr and all(
                                            w in rname_lc for w in words
                                        ):
                                            room_id = rid
                                            break
                            if room_id:
                                via_location += 1

                    if room_id:
                        device.room_id = room_id
                        if room_id not in line_room_ids:
                            line_room_ids.append(room_id)

                for room_id in line_room_ids:
                    if room_id not in line.assigned_room_ids:
                        line.assigned_room_ids.append(room_id)
                        linked_pairs += 1

        logger.info(
            f"link_rooms_to_lines: {linked_pairs} Paare — "
            f"{via_note} via Installations-Hinweis, {via_ga} via GA-Bezeichnung, "
            f"{via_location} via Einbauort."
        )
        return linked_pairs

    def enrich_device_ko_connections(
        self,
        topology: Topology,
        filepath: str,
    ) -> tuple[int, int]:
        """
        Ergänzt KO-Verbindungen in der Topologie aus einem ETS6 GA-Report (XLSX).

        Der GA-Report enthält pro Gerät alle Kanäle mit vollständiger GA-Liste
        (col 28 = alle verbundenen GAs, nicht nur die aktuelle).
        Diese Methode:
          - ergänzt fehlende connected_gas-Einträge in bestehenden KOs
          - fügt vollständig fehlende KOs hinzu

        Gibt (enriched_devices, added_cos) zurück.
        """
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"XLSX-Datei nicht gefunden: {filepath}")
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl wird für XLSX-Import benoetigt.")

        # ── Schritt 1: GA-Report parsen → {phys_addr → {ko_num → CO}} ──
        parsed: dict[str, dict[int, CommunicationObject]] = {}
        # Gerätemetadaten aus GA-Report (Produkt, Hersteller) für neue Geräte
        parsed_device_meta: dict[str, dict] = {}

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        current_phys: Optional[str] = None

        for row in ws.iter_rows(values_only=True):
            if not row or COL_GAR_ADDR >= len(row) or row[COL_GAR_ADDR] is None:
                continue
            addr_raw = str(row[COL_GAR_ADDR]).strip()

            # Gerätezeile
            if _PHYS_ADDR_RE.match(addr_raw):
                current_phys = addr_raw
                if current_phys not in parsed:
                    parsed[current_phys] = {}
                    parsed_device_meta[current_phys] = {
                        "product": _cell(row, COL_GAR_PRODUCT),
                        "location": _cell(row, COL_GAR_LOCATION),
                    }
                continue

            # KO-Zeile
            if current_phys and _GAR_KO_RE.match(addr_raw):
                m_ko = re.match(r"^(\d+):\s*(.+)$", addr_raw)
                if not m_ko:
                    continue
                ko_num = int(m_ko.group(1))
                ko_desc = m_ko.group(2).strip()

                # KO-Name und Funktion trennen: "Ausgang B - Schalten" / "G1, Schalten, - An/Aus"
                parts = ko_desc.rsplit(" - ", 1)
                ko_name = parts[0].strip()
                ko_func = parts[1].strip() if len(parts) == 2 else ""

                gas_raw = row[COL_GAR_KO_GAS] if COL_GAR_KO_GAS < len(row) else None
                connected_gas = _GA_RE.findall(str(gas_raw)) if gas_raw else []

                if ko_num not in parsed[current_phys]:
                    parsed[current_phys][ko_num] = CommunicationObject(
                        object_number=ko_num,
                        name=ko_name,
                        object_function=ko_func,
                        data_type=_cell(row, COL_GAR_KO_DTYPE),
                        priority=_cell(row, COL_GAR_KO_PRIO) or "Niedrig",
                        flags=_cell(row, COL_GAR_KO_FLAGS),
                        connected_gas=connected_gas,
                    )
                else:
                    # KO bereits vorhanden: fehlende GAs ergänzen
                    existing = parsed[current_phys][ko_num]
                    for ga in connected_gas:
                        if ga not in existing.connected_gas:
                            existing.connected_gas.append(ga)

            # Neue GA-Zeile → nächste Gerätesektion beginnt beim nächsten Gerät
            elif _GAR_GA_RE.match(addr_raw):
                current_phys = None

        wb.close()

        # ── Schritt 2: Topologie-Geräte anreichern ──
        enriched_devices = 0
        added_cos = 0
        added_devices = 0

        # Index: phys_addr → Device
        device_index: dict[str, Device] = {
            dev.physical_address: dev
            for area in topology.areas
            for line in area.lines
            for dev in line.devices
            if dev.physical_address
        }

        # Index: line_addr (z.B. "1.1") → Line
        line_index: dict[str, Line] = {
            f"{area.area_number}.{line.line_number}": line
            for area in topology.areas
            for line in area.lines
        }

        for phys_addr, ko_map in parsed.items():
            device = device_index.get(phys_addr)
            if not device:
                # Gerät nur im GA-Report vorhanden (z.B. Taster ohne Topologie-Eintrag)
                # → anhand der physikalischen Adresse der richtigen Linie zuordnen
                m = _PHYS_ADDR_RE.match(phys_addr)
                if not m:
                    continue
                line_addr = f"{m.group(1)}.{m.group(2)}"
                target_line = line_index.get(line_addr)
                if not target_line:
                    continue
                meta = parsed_device_meta.get(phys_addr, {})
                device = Device(
                    physical_address=phys_addr,
                    product=meta.get("product", ""),
                    installation_location=meta.get("location", ""),
                    device_type=self._classify_device(phys_addr, meta.get("product", "")),
                )
                target_line.devices.append(device)
                device_index[phys_addr] = device
                added_devices += 1

            # Index bestehender KOs: ko_num → CO
            existing_kos: dict[int, CommunicationObject] = {
                co.object_number: co for co in device.communication_objects
            }

            changed = False
            for ko_num, new_co in ko_map.items():
                if ko_num in existing_kos:
                    # Fehlende GAs in bestehendem KO ergänzen
                    existing = existing_kos[ko_num]
                    for ga in new_co.connected_gas:
                        if ga not in existing.connected_gas:
                            existing.connected_gas.append(ga)
                            changed = True
                else:
                    # KO komplett neu hinzufügen
                    device.communication_objects.append(new_co)
                    added_cos += 1
                    changed = True

            if changed:
                enriched_devices += 1

        logger.info(
            f"enrich_device_ko_connections: {enriched_devices} Geräte angereichert, "
            f"{added_cos} neue KOs hinzugefügt, "
            f"{added_devices} fehlende Geräte (Taster/Sensor) ergänzt  [{filepath}]"
        )
        if added_devices:
            for area in topology.areas:
                for line in area.lines:
                    line.update_device_count()
        return enriched_devices, added_cos
