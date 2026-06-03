"""
Excel-Erzeugung (NFA-050)
Basis-Excel-Engine für Berichte und Formulare mit openpyxl.
"""
from __future__ import annotations
import logging
from datetime import datetime

logger = logging.getLogger("knix_arranger.excel_generator")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.info("openpyxl nicht installiert. Excel-Export nicht verfuegbar.")

# KNX-Farben
KNX_GREEN_HEX      = "5EA126"
KNX_DARK_GREEN_HEX = "3A6B19"
KNX_BLUE_HEX       = "1A5276"
KNX_MID_BLUE_HEX   = "2E86C1"
KNX_LIGHT_BLUE_HEX = "D6EAF8"
WHITE_HEX           = "FFFFFF"
LIGHT_GRAY_HEX      = "F2F2F2"
MEDIUM_GRAY_HEX     = "E8E8E8"
RED_HEX             = "E74C3C"
YELLOW_HEX          = "F39C12"

_THIN_BORDER_SIDE = None   # lazy init


def _thin_border():
    global _THIN_BORDER_SIDE
    if _THIN_BORDER_SIDE is None:
        s = Side(style="thin")
        _THIN_BORDER_SIDE = Border(left=s, right=s, top=s, bottom=s)
    return _THIN_BORDER_SIDE


class ExcelGenerator:
    """Erzeugt Excel-Dokumente für Berichte und Formulare."""

    def __init__(self, title: str = "", company: str = "",
                 project_name: str = ""):
        if not HAS_OPENPYXL:
            raise ImportError(
                "openpyxl ist nicht installiert. "
                "Bitte installieren mit: pip install openpyxl"
            )
        self.title = title
        self.company = company
        self.project_name = project_name
        self.wb = Workbook()
        self._current_sheet = self.wb.active
        self._current_sheet.title = title[:31] if title else "Bericht"
        self._row = 1
        self._num_cols = 6   # Standard-Spaltenbreite für Merge

    def add_header(self):
        """Fügt einen Berichts-Header hinzu."""
        ws = self._current_sheet
        nc = self._num_cols

        header_font = Font(name="Arial", size=14, bold=True, color=WHITE_HEX)
        header_fill = PatternFill(start_color=KNX_GREEN_HEX,
                                  end_color=KNX_GREEN_HEX, fill_type="solid")
        ws.merge_cells(start_row=self._row, start_column=1,
                       end_row=self._row, end_column=nc)
        cell = ws.cell(row=self._row, column=1, value=self.title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   indent=1)
        ws.row_dimensions[self._row].height = 26
        self._row += 1

        info_font = Font(name="Arial", size=9, italic=True,
                         color=KNX_DARK_GREEN_HEX)
        info_fill = PatternFill(start_color=MEDIUM_GRAY_HEX,
                                end_color=MEDIUM_GRAY_HEX, fill_type="solid")
        ws.merge_cells(start_row=self._row, start_column=1,
                       end_row=self._row, end_column=nc)
        info_text = (
            f"Projekt: {self.project_name}    |    "
            f"Datum: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            + (f"    |    Firma: {self.company}" if self.company else "")
        )
        cell = ws.cell(row=self._row, column=1, value=info_text)
        cell.font = info_font
        cell.fill = info_fill
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                   indent=1)
        ws.row_dimensions[self._row].height = 16
        self._row += 2

    def add_heading(self, text: str, level: int = 1):
        """Fügt eine Überschrift mit level-abhängigem Styling hinzu.

        Level 1 – Dokumenttitel (grün, weiß)
        Level 2 – Raumüberschrift (blau, weiß)
        Level 3 – Unterabschnitt/Bedienelement (hellblau, dunkel)
        """
        ws = self._current_sheet
        nc = self._num_cols

        if level == 1:
            font_color, fill_color = WHITE_HEX, KNX_GREEN_HEX
            font_size, row_h, indent = 13, 22, 1
        elif level == 2:
            font_color, fill_color = WHITE_HEX, KNX_MID_BLUE_HEX
            font_size, row_h, indent = 11, 20, 1
        else:
            font_color, fill_color = KNX_DARK_GREEN_HEX, KNX_LIGHT_BLUE_HEX
            font_size, row_h, indent = 10, 17, 2

        font = Font(name="Arial", size=font_size, bold=True, color=font_color)
        fill = PatternFill(start_color=fill_color, end_color=fill_color,
                           fill_type="solid")
        ws.merge_cells(start_row=self._row, start_column=1,
                       end_row=self._row, end_column=nc)
        cell = ws.cell(row=self._row, column=1, value=text)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", indent=indent)
        ws.row_dimensions[self._row].height = row_h
        self._row += 1

    def add_paragraph(self, text: str):
        """Fügt einen Textabsatz hinzu."""
        ws = self._current_sheet
        ws.merge_cells(start_row=self._row, start_column=1,
                       end_row=self._row, end_column=self._num_cols)
        cell = ws.cell(row=self._row, column=1, value=text)
        cell.font = Font(name="Arial", size=9, italic=True,
                         color=KNX_DARK_GREEN_HEX)
        cell.alignment = Alignment(indent=2, vertical="center")
        ws.row_dimensions[self._row].height = 15
        self._row += 1

    def add_empty_row(self, height: int = 6):
        ws = self._current_sheet
        ws.row_dimensions[self._row].height = height
        self._row += 1

    def set_column_widths(self, widths: list[float]):
        """Setzt Spaltenbreiten für das aktuelle Tabellenblatt (einmalig aufrufen)."""
        ws = self._current_sheet
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        self._num_cols = max(len(widths), self._num_cols)

    def add_table(self, headers: list[str], rows: list[list],
                  col_widths: list[int] | None = None,
                  center_cols: list[int] | None = None,
                  data_row_height: int = 22,
                  wrap_cols: list[int] | None = None):
        """Fügt eine formatierte Tabelle hinzu.

        center_cols: 1-basierte Spaltennummern, die zentriert werden sollen.
        wrap_cols:   1-basierte Spaltennummern mit Zeilenumbruch.
        """
        ws = self._current_sheet
        border = _thin_border()

        header_font = Font(name="Arial", size=9, bold=True, color=WHITE_HEX)
        header_fill = PatternFill(start_color=KNX_BLUE_HEX,
                                  end_color=KNX_BLUE_HEX, fill_type="solid")
        row_font = Font(name="Arial", size=9)
        alt_fill = PatternFill(start_color=LIGHT_GRAY_HEX,
                               end_color=LIGHT_GRAY_HEX, fill_type="solid")

        if col_widths:
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

        center_set = set(center_cols or [])
        wrap_set   = set(wrap_cols or [])

        # Header-Zeile
        ws.row_dimensions[self._row].height = 16
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=self._row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
        self._row += 1

        # Daten-Zeilen
        for row_idx, row_data in enumerate(rows):
            ws.row_dimensions[self._row].height = data_row_height
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=self._row, column=col_idx, value=value)
                cell.font = row_font
                cell.border = border
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
                h_align = "center" if col_idx in center_set else "left"
                cell.alignment = Alignment(
                    horizontal=h_align,
                    vertical="center",
                    wrap_text=(col_idx in wrap_set),
                    indent=0 if col_idx in center_set else 1,
                )
            self._row += 1

        self._row += 1   # Abstand nach Tabelle

    def set_print_options(self, orientation: str = "landscape",
                          paper_size: int = 9,    # 9 = A4
                          fit_to_page: bool = True):
        """Konfiguriert Druckoptionen für das aktuelle Tabellenblatt."""
        ws = self._current_sheet
        ws.page_setup.orientation = orientation
        ws.page_setup.paperSize = paper_size
        if fit_to_page:
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
        ws.print_options.gridLines = False
        ws.print_options.headings = False

    def add_sheet(self, name: str):
        """Fügt ein neues Tabellenblatt hinzu."""
        self._current_sheet = self.wb.create_sheet(title=name[:31])
        self._row = 1

    def save(self, filepath: str):
        """Speichert die Excel-Datei."""
        if not filepath.endswith((".xlsx", ".xls")):
            filepath += ".xlsx"
        self.wb.save(filepath)
        logger.info(f"Excel gespeichert: {filepath}")
